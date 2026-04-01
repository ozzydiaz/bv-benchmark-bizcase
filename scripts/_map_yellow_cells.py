"""
Map yellow-highlighted input cells to their row labels and section headers.
Outputs a structured view of every user-editable field in the source workbook.
"""
import openpyxl

wb = openpyxl.load_workbook(
    "Template_BV Benchmark Business Case v6.xlsm",
    keep_vba=True,
    data_only=True,
)


def is_yellow(color_str: str) -> bool:
    if not color_str:
        return False
    c = color_str.upper().lstrip("#")
    if len(c) == 8:
        r, g, b = int(c[2:4], 16), int(c[4:6], 16), int(c[6:8], 16)
    elif len(c) == 6:
        r, g, b = int(c[0:2], 16), int(c[2:4], 16), int(c[4:6], 16)
    else:
        return False
    return r > 180 and g > 180 and b < 130


def get_row_label(ws, row_num):
    """Try columns B and C for a non-empty label."""
    for col in range(1, 6):
        v = ws.cell(row=row_num, column=col).value
        if v and str(v).strip():
            return str(v).strip()
    return ""


def scan_sheet(sheet_name):
    ws = wb[sheet_name]
    print(f"\n{'='*70}")
    print(f"  {sheet_name}")
    print(f"{'='*70}")
    current_section = ""
    for row in ws.iter_rows():
        for cell in row:
            fill = cell.fill
            if not fill or fill.fill_type in (None, "none"):
                continue
            fg = fill.fgColor
            color = None
            if fg.type == "rgb":
                color = fg.rgb
            if color and is_yellow(color):
                label = get_row_label(ws, cell.row)
                print(f"  {cell.coordinate:<8}  {repr(cell.value):<30}  {label}")


scan_sheet("1-Client Variables")
scan_sheet("2a-Consumption Plan Wk1")
