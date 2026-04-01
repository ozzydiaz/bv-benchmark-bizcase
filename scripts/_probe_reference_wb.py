"""Probe the filled reference workbook output cells."""
import openpyxl
wb = openpyxl.load_workbook(
    "Reliance_BV Benchmark Business Case v6.xlsm", keep_vba=True, data_only=True
)

sfc = wb["Summary Financial Case"]
print("=== Summary Financial Case (all non-None rows 1-80) ===")
for row in sfc.iter_rows(min_row=1, max_row=80, max_col=12):
    for cell in row:
        if cell.value is not None:
            print(f"  {cell.coordinate:<8} {repr(cell.value)[:70]}")

cf5 = wb["5Y CF with Payback"]
print("\n=== 5Y CF with Payback (rows 1-55) ===")
for row in cf5.iter_rows(min_row=1, max_row=55, max_col=12):
    for cell in row:
        if cell.value is not None:
            print(f"  {cell.coordinate:<8} {repr(cell.value)[:70]}")
