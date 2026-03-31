"""Probe RVtools columns and print sample data for verification."""
import openpyxl

RVTOOLS_FILE = "RVTools_export_VCP003_2026-01-05_13.14.03.xlsx"

wb = openpyxl.load_workbook(RVTOOLS_FILE, data_only=True, read_only=True)

ws = wb['vInfo']
rows = list(ws.iter_rows(max_row=5, values_only=True))
headers = list(rows[0])
key_names = ['VM', 'Powerstate', 'CPUs', 'Memory', 'In Use MiB',
             'OS according to the configuration file',
             'OS according to the VMware Tools']
print("=== vInfo key columns ===")
for name in key_names:
    idx = headers.index(name) if name in headers else None
    sample = rows[1][idx] if idx is not None and len(rows) > 1 else "NOT FOUND"
    print(f"  col {str(idx):>4}: {name:50s} sample={sample}")

ws2 = wb['vHost']
rows2 = list(ws2.iter_rows(max_row=3, values_only=True))
headers2 = list(rows2[0])
host_names = ['Host', '# Cores', '# Memory', 'vCPUs per Core', '# VMs total']
print("\n=== vHost key columns ===")
for name in host_names:
    idx = headers2.index(name) if name in headers2 else None
    sample = rows2[1][idx] if idx is not None and len(rows2) > 1 else "NOT FOUND"
    print(f"  col {str(idx):>4}: {name:50s} sample={sample}")
