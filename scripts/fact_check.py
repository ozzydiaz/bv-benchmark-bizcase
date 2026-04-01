"""
scripts/fact_check.py
=====================
CLI wrapper for the engine fact checker.

Usage:
    python scripts/fact_check.py --workbook path/to/client.xlsm [--strict]

The workbook must have been saved in Excel (so formula cells contain cached
numeric values).  The script reads the yellow input cells from the workbook
to reconstruct BusinessCaseInputs automatically, then runs the engine and
compares outputs against the Excel-cached values.

Exit codes:
    0 — all critical checks passed (WARN is allowed)
    1 — one or more FAIL checks; use --strict to also fail on WARN
"""
from __future__ import annotations

import argparse
import pathlib
import sys

import openpyxl

# Allow running from repo root without installing
sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))

from engine.models import (
    BusinessCaseInputs, EngagementInfo, PricingConfig, DatacenterConfig,
    HardwareLifecycle, WorkloadInventory, ConsumptionPlan, AzureRunRate, YesNo,
    BenchmarkConfig,
)
from engine.fact_checker import run as fact_check_run, CLIENT_VAR_CELLS, CONSUMPTION_CELLS

_REPO_ROOT = pathlib.Path(__file__).parent.parent
_BENCHMARKS_YAML = _REPO_ROOT / "data" / "benchmarks_default.yaml"


def _cell_float(ws, addr: str, default: float = 0.0) -> float:
    v = ws[addr].value
    try:
        return float(v) if v is not None else default
    except (TypeError, ValueError):
        return default


def _cell_str(ws, addr: str, default: str = "") -> str:
    v = ws[addr].value
    return str(v).strip() if v is not None else default


def _read_optional_int(ws, addr: str) -> int | None:
    """Return int from a cell, or None if empty (so model auto-derive kicks in)."""
    v = ws[addr].value
    if v is None:
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def build_inputs_from_workbook(wb: openpyxl.Workbook) -> BusinessCaseInputs:
    """
    Reconstruct BusinessCaseInputs by reading the yellow input cells from
    the workbook's '1-Client Variables' and '2a-Consumption Plan Wk1' sheets.
    """
    cv = wb["1-Client Variables"]
    cp_ws = wb["2a-Consumption Plan Wk1"]

    # Engagement
    client_name = _cell_str(cv, CLIENT_VAR_CELLS["client_name"], "Client")
    currency = _cell_str(cv, CLIENT_VAR_CELLS["local_currency"], "USD")

    # Workload inventory (Workload #1)
    wl = WorkloadInventory(
        workload_name=_cell_str(cv, CLIENT_VAR_CELLS["workload_name"], "DC Move"),
        num_vms=int(_cell_float(cv, CLIENT_VAR_CELLS["num_vms"])),
        num_physical_servers_excl_hosts=int(_cell_float(cv, CLIENT_VAR_CELLS["num_physical_servers"])),
        allocated_vcpu=int(_cell_float(cv, CLIENT_VAR_CELLS["allocated_vcpu"])),
        allocated_pcores_excl_hosts=int(_cell_float(cv, CLIENT_VAR_CELLS["allocated_pcores"])),
        allocated_vmemory_gb=_cell_float(cv, CLIENT_VAR_CELLS["allocated_vmemory_gb"]),
        allocated_pmemory_gb_excl_hosts=_cell_float(cv, CLIENT_VAR_CELLS["allocated_pmemory_gb"]),
        allocated_storage_gb=_cell_float(cv, CLIENT_VAR_CELLS["allocated_storage_gb"]),
        vcpu_per_core_ratio=_cell_float(cv, CLIENT_VAR_CELLS["vcpu_per_core_ratio"], 1.97),
        pcores_with_windows_server=int(_cell_float(cv, CLIENT_VAR_CELLS["pcores_windows_server"])),
        pcores_with_windows_esu=int(_cell_float(cv, CLIENT_VAR_CELLS["pcores_windows_esu"])),
        # D70/D71: formula defaults to 10% of D67/D68 in Template; overridden with
        # actuals when available from RVtools or OS audit. Read directly so the
        # fact checker uses the same values as the workbook — not the Python auto-derive.
        pcores_with_sql_server=_read_optional_int(cv, CLIENT_VAR_CELLS["pcores_sql_server"]),
        pcores_with_sql_esu=_read_optional_int(cv, CLIENT_VAR_CELLS["pcores_sql_esu"]),
        # D58–D61: backup/DR sizing — only meaningful when the corresponding option
        # is activated AND included in the Azure consumption plan. If blank in the
        # workbook, pass None so the engine treats backup/DR storage cost as 0.
        backup_size_gb=_cell_float(cv, CLIENT_VAR_CELLS["backup_size_gb"]) or None,
        backup_num_protected_vms=_read_optional_int(cv, CLIENT_VAR_CELLS["backup_num_protected_vms"]),
        dr_size_gb=_cell_float(cv, CLIENT_VAR_CELLS["dr_size_gb"]) or None,
        dr_num_protected_vms=_read_optional_int(cv, CLIENT_VAR_CELLS["dr_num_protected_vms"]),
    )

    # Consumption plan (Workload #1)
    ramp_addrs = ["E17","F17","G17","H17","I17","J17","K17","L17","M17","N17"]
    ramp = [_cell_float(cp_ws, a, 1.0) for a in ramp_addrs]

    aco_addrs  = ["E21","F21","G21","H21","I21","J21","K21","L21","M21","N21"]
    ecif_addrs = ["E22","F22","G22","H22","I22","J22","K22","L22","M22","N22"]
    aco  = [_cell_float(cp_ws, a) for a in aco_addrs]
    ecif = [_cell_float(cp_ws, a) for a in ecif_addrs]

    # Compute consumption: Y9 (col M) is the steady-state anchor
    compute_y10 = _cell_float(cp_ws, "M28")
    storage_y10 = _cell_float(cp_ws, "M29")
    other_y10   = _cell_float(cp_ws, "M30")

    backup_on = _cell_str(cp_ws, CONSUMPTION_CELLS["backup_activated"], "No")
    backup_stor = _cell_str(cp_ws, CONSUMPTION_CELLS["backup_stor_in_consumption"], "No")
    backup_sw = _cell_str(cp_ws, CONSUMPTION_CELLS["backup_sw_in_consumption"], "No")
    dr_on     = _cell_str(cp_ws, CONSUMPTION_CELLS["dr_activated"], "No")
    dr_stor   = _cell_str(cp_ws, CONSUMPTION_CELLS["dr_stor_in_consumption"], "No")
    dr_sw     = _cell_str(cp_ws, CONSUMPTION_CELLS["dr_sw_in_consumption"], "No")

    cp = ConsumptionPlan(
        workload_name=wl.workload_name,
        azure_vcpu=int(_cell_float(cp_ws, CONSUMPTION_CELLS["azure_vcpu"])),
        azure_memory_gb=_cell_float(cp_ws, CONSUMPTION_CELLS["azure_memory_gb"]),
        azure_storage_gb=_cell_float(cp_ws, CONSUMPTION_CELLS["azure_storage_gb"]),
        migration_ramp_pct=ramp,
        aco_by_year=aco,
        ecif_by_year=ecif,
        annual_compute_consumption_lc_y10=compute_y10,
        annual_storage_consumption_lc_y10=storage_y10,
        annual_other_consumption_lc_y10=other_y10,
        backup_activated=YesNo(backup_on) if backup_on in ("Yes", "No") else YesNo.NO,
        backup_storage_in_consumption=YesNo(backup_stor) if backup_stor in ("Yes", "No") else YesNo.NO,
        backup_software_in_consumption=YesNo(backup_sw) if backup_sw in ("Yes", "No") else YesNo.NO,
        dr_activated=YesNo(dr_on) if dr_on in ("Yes", "No") else YesNo.NO,
        dr_storage_in_consumption=YesNo(dr_stor) if dr_stor in ("Yes", "No") else YesNo.NO,
        dr_software_in_consumption=YesNo(dr_sw) if dr_sw in ("Yes", "No") else YesNo.NO,
    )

    # Azure Run Rate
    rr_include = _cell_str(cv, CLIENT_VAR_CELLS["include_run_rate"], "No")
    arr = AzureRunRate(
        include_in_business_case=YesNo(rr_include) if rr_include in ("Yes", "No") else YesNo.NO,
        current_acd=_cell_float(cv, CLIENT_VAR_CELLS["current_acd"]),
        new_acd=_cell_float(cv, CLIENT_VAR_CELLS["new_acd"]),
        monthly_spend_usd=_cell_float(cv, CLIENT_VAR_CELLS["monthly_spend_usd"]),
        paygo_mix=_cell_float(cv, CLIENT_VAR_CELLS["paygo_mix"], 1.0),
        reserved_instances_mix=_cell_float(cv, CLIENT_VAR_CELLS["ri_mix"]),
        savings_plan_mix=_cell_float(cv, CLIENT_VAR_CELLS["sp_mix"]),
        sku_discount_mix=_cell_float(cv, CLIENT_VAR_CELLS["sku_mix"]),
    )

    # Hardware lifecycle — read D26 for growth rate; other fields use model defaults
    growth_rate = _cell_float(cv, "D26", 0.10)
    hardware = HardwareLifecycle(expected_future_growth_rate=growth_rate)

    return BusinessCaseInputs(
        engagement=EngagementInfo(
            client_name=client_name,
            local_currency_name=currency,
        ),
        hardware=hardware,
        workloads=[wl],
        consumption_plans=[cp],
        azure_run_rate=arr,
    )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Fact-check the Python engine against a saved client workbook."
    )
    parser.add_argument(
        "--workbook", "-w", required=True,
        help="Path to a saved .xlsm/.xlsx workbook with cached formula values.",
    )
    parser.add_argument(
        "--strict", action="store_true",
        help="Exit code 1 on WARN checks as well as FAIL.",
    )
    parser.add_argument(
        "--json", action="store_true",
        help="Output results as JSON (useful for CI pipelines).",
    )
    args = parser.parse_args()

    wb_path = pathlib.Path(args.workbook)
    if not wb_path.exists():
        print(f"ERROR: workbook not found: {wb_path}", file=sys.stderr)
        return 2

    print(f"Loading workbook: {wb_path.name} ...")
    wb = openpyxl.load_workbook(str(wb_path), keep_vba=True, data_only=True)

    print("Reconstructing inputs from workbook ...")
    inputs = build_inputs_from_workbook(wb)

    benchmarks = BenchmarkConfig.from_yaml(str(_BENCHMARKS_YAML))

    print("Running fact check ...\n")
    report = fact_check_run(str(wb_path), inputs, benchmarks)
    report.print()

    if args.json:
        import json
        out = {
            "confidence_score": report.confidence_score,
            "passed": report.passed,
            "warned": report.warned,
            "failed": report.failed,
            "skipped": report.skipped,
            "input_mismatches": report.input_mismatches,
            "checks": [
                {
                    "name": c.name,
                    "excel": c.excel_value,
                    "engine": c.engine_value,
                    "delta_pct": c.delta_pct,
                    "status": c.status,
                }
                for c in report.checks
            ],
        }
        print(json.dumps(out, indent=2))

    if report.failed > 0:
        return 1
    if args.strict and report.warned > 0:
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
