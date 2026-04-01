"""Scan 1-Client Variables and 2a-Consumption Plan Wk1 for yellow-highlighted input cells."""
import openpyxl

wb = openpyxl.load_workbook(
    "Template_BV Benchmark Business Case v6.xlsm",
    keep_vba=True,
    data_only=True,
)


def is_yellow(color_str: str) -> bool:
    if not color_str:
        return False
    c = color_str.upper().lstrip("#")
    if len(c) == 8:
        r, g, b = int(c[2:4], 16), int(c[4:6], 16), int(c[6:8], 16)
    elif len(c) == 6:
        r, g, b = int(c[0:2], 16), int(c[2:4], 16), int(c[4:6], 16)
    else:
        return False
    # Yellow: high red, high green, low blue
    return r > 180 and g > 180 and b < 130


for sheet_name in ["1-Client Variables", "2a-Consumption Plan Wk1"]:
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ===")
    for row in ws.iter_rows():
        for cell in row:
            fill = cell.fill
            if not fill or fill.fill_type in (None, "none"):
                continue
            fg = fill.fgColor
            color = None
            if fg.type == "rgb":
                color = fg.rgb
            if color and is_yellow(color):
                # Try to get label from column B or adjacent
                label_cell = ws.cell(row=cell.row, column=2)
                label = label_cell.value if label_cell.value else ""
                print(f"  {cell.coordinate:8s}  color={color}  label={repr(label)[:50]:50s}  value={repr(cell.value)}")
