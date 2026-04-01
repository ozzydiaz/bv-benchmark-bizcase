"""Rationale probe — shows the exact distributions, math, and assumptions behind each output."""
import sys, math
sys.path.insert(0, '.')
import openpyxl
import numpy as np

WB = 'RVTools_export_VCP003_2026-01-05_13.14.03.xlsx'
wb = openpyxl.load_workbook(WB, data_only=True, read_only=True)

# ─── 1. CPU utilisation distribution ─────────────────────────────────────────
ws = wb['vCPU']
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
ci_ps  = headers.index('Powerstate')
ci_max = headers.index('Max')
ci_ovr = headers.index('Overall')

cpu_utils = []
for row in rows[1:]:
    if str(row[ci_ps] or '').lower() != 'poweredon': continue
    mx, ov = row[ci_max], row[ci_ovr]
    if isinstance(mx,(int,float)) and isinstance(ov,(int,float)) and mx > 0:
        cpu_utils.append(ov/mx)

cpu_utils.sort()
n = len(cpu_utils)
p50 = cpu_utils[int(n*0.50)]
p75 = cpu_utils[int(n*0.75)]
p90 = cpu_utils[int(n*0.90)]
p95 = cpu_utils[int(n*0.95)]
p99 = cpu_utils[int(n*0.99)]
print(f'CPU utilisation — {n:,} powered-on VMs with valid MHz readings')
print(f'  P50 = {p50:.1%}   P75 = {p75:.1%}   P90 = {p90:.1%}   P95 = {p95:.1%}   P99 = {p99:.1%}')
print(f'  Mean = {sum(cpu_utils)/n:.1%}')
zero_or_idle = sum(1 for x in cpu_utils if x == 0.0)
over_50 = sum(1 for x in cpu_utils if x > 0.50)
print(f'  VMs at 0% utilisation (idle/no CPU work at snapshot): {zero_or_idle:,} ({zero_or_idle/n:.0%})')
print(f'  VMs above 50%: {over_50:,} ({over_50/n:.0%})')

base_vcpu = 8328
azure_vcpu = math.ceil(base_vcpu * p95 * 1.20)
print(f'\n  Calculation: ceil({base_vcpu:,} × {p95:.4f} × 1.20) = ceil({base_vcpu * p95 * 1.20:.1f}) = {azure_vcpu:,}')
print(f'  Reduction: {(1 - azure_vcpu/base_vcpu):.0%}')

# ─── 2. Memory utilisation distribution ──────────────────────────────────────
ws2 = wb['vMemory']
rows2 = list(ws2.iter_rows(values_only=True))
headers2 = rows2[0]
ci_ps2  = headers2.index('Powerstate')
ci_sz   = headers2.index('Size MiB')
ci_con  = headers2.index('Consumed')
ci_bal  = headers2.index('Ballooned')

mem_utils = []
ballooned_vms = 0
for row in rows2[1:]:
    if str(row[ci_ps2] or '').lower() != 'poweredon': continue
    sz, co = row[ci_sz], row[ci_con]
    bal = row[ci_bal] or 0
    if isinstance(sz,(int,float)) and isinstance(co,(int,float)) and sz > 0:
        mem_utils.append(co/sz)
        if isinstance(bal,(int,float)) and bal > 0:
            ballooned_vms += 1

mem_utils.sort()
nm = len(mem_utils)
mp50 = mem_utils[int(nm*0.50)]
mp75 = mem_utils[int(nm*0.75)]
mp90 = mem_utils[int(nm*0.90)]
mp95 = mem_utils[int(nm*0.95)]
mp99 = mem_utils[int(nm*0.99)]
print(f'\nMemory utilisation — {nm:,} powered-on VMs')
print(f'  P50 = {mp50:.1%}   P75 = {mp75:.1%}   P90 = {mp90:.1%}   P95 = {mp95:.1%}   P99 = {mp99:.1%}')
print(f'  Mean = {sum(mem_utils)/nm:.1%}')
over_80 = sum(1 for x in mem_utils if x > 0.80)
over_100 = sum(1 for x in mem_utils if x >= 1.0)
print(f'  VMs above 80% consumed: {over_80:,} ({over_80/nm:.0%})')
print(f'  VMs at/above 100% consumed: {over_100:,} ({over_100/nm:.0%})')
print(f'  VMs with active ballooning: {ballooned_vms:,} ({ballooned_vms/nm:.0%})')
base_mem = 31115.5
azure_mem = math.ceil(base_mem * mp95 * 1.20)
print(f'\n  Calculation: ceil({base_mem:,.1f} × {mp95:.4f} × 1.20) = ceil({base_mem * mp95 * 1.20:.1f}) = {azure_mem:,} GB')

# ─── 3. Region evidence ───────────────────────────────────────────────────────
ws3 = wb['vHost']
rows3 = list(ws3.iter_rows(values_only=True))
headers3 = rows3[0]
ci_dc  = headers3.index('Datacenter')
ci_tz  = headers3.index('Time Zone Name')
ci_gmt = headers3.index('GMT Offset')
ci_dom = headers3.index('Domain')
ci_h   = headers3.index('Host')

dcs, tzs, gmts, doms, hosts = set(), set(), set(), set(), []
for row in rows3[1:]:
    if row[ci_h] is None: continue
    if row[ci_dc]:  dcs.add(str(row[ci_dc]).strip())
    if row[ci_tz]:  tzs.add(str(row[ci_tz]).strip())
    if row[ci_gmt] is not None: gmts.add(str(row[ci_gmt]).strip())
    if row[ci_dom]: doms.add(str(row[ci_dom]).strip().lower())
    hosts.append(str(row[ci_h]))

print(f'\nRegion evidence from vHost ({len(hosts)} hosts):')
print(f'  Datacenter names : {sorted(dcs)}')
print(f'  Time zones       : {sorted(tzs)}')
print(f'  GMT offsets      : {sorted(gmts)}   <- minutes east of UTC')
print(f'  Domain names     : {sorted(doms)}')
print(f'  Host FQDNs (sample): {hosts[:3]}')

ws4 = wb['vMetaData']
rows4 = list(ws4.iter_rows(values_only=True))
headers4 = rows4[0]
ci_srv = headers4.index('Server')
fqdns = [str(row[ci_srv]) for row in rows4[1:] if row[ci_srv]]
print(f'  vCenter FQDNs    : {fqdns}')

# ─── 4. Cost math ─────────────────────────────────────────────────────────────
print(f'\nCost calculations:')
vcpu_rate = 0.048
gb_rate   = 0.018
hrs       = 8760
compute   = azure_vcpu * hrs * vcpu_rate
stor_gb   = math.ceil(1185982.1 * 1.20)
storage   = stor_gb * 12 * gb_rate
print(f'  Compute: {azure_vcpu:,} vCPU × {hrs:,} hr × ${vcpu_rate}/vCPU/hr = ${compute:,.0f}/yr')
print(f'  Storage: {stor_gb:,} GB × 12 mo × ${gb_rate}/GB/mo = ${storage:,.0f}/yr')
print(f'  Benchmark rate ${vcpu_rate}/vCPU/hr source: Standard_D4s_v5 East US PAYG = $0.192/hr ÷ 4 vCPU')
print(f'  Benchmark rate ${gb_rate}/GB/mo source: Standard SSD E10 LRS East US $2.304/mo ÷ 128 GiB')
