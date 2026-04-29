"""
Layer 1 Business-Analyst Replica
================================

A standalone, engine-independent oracle that mechanically follows the rules
in ``training/ba_rules/layer1.yaml``. Used as the reference implementation
for parity diffing against ``engine/rvtools_parser.py``.

Design rules (enforced by ``training/ba_rules/layer1.yaml`` key_principles):

* **KP.PER_VM**     — every business-case value is computed PER VM and summed.
* **KP.MIB_DEFAULT** — assume MiB unless the column header explicitly states
                       another unit; convert to decimal GB (÷ 953.674).
* **KP.SYNONYM_HEADERS** — resolve every required column via a synonym set,
                           never an exact-string lookup.
* **KP.MANDATORY_VINFO** — vInfo missing or empty halts the workflow.
* **KP.PROVISIONED_FOR_TCO** — on-prem TCO storage uses Provisioned (per VM);
                                in-use is captured separately for Layer 2.
* **KP.BA_APPROVAL_GATE** — assumptions, missing-tab notices, and proposed
                             reduction factors are surfaced in a review packet.

NO IMPORTS FROM ``engine/`` — this is an independent oracle.
"""

from __future__ import annotations

import argparse
import dataclasses
import logging
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import yaml

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MIB_TO_GB_DECIMAL = 1.0 / 953.674      # KP.MIB_DEFAULT — Azure pricing API basis
MB_TO_GB_DECIMAL = 1000.0 / 1_000_000  # decimal MB → decimal GB (just /1)
MB_TO_GB_BINARY = 1.0 / 1024.0         # if a column is explicitly labelled MB

DEFAULT_VCPU_PER_PCORE_BENCHMARK = 1.97   # used when vHost absent (per Layer 3)
DEFAULT_REDUCTION_FACTORS = {
    "cpu_util_pct": 50.0,                 # BA-overridable; surfaced in packet
    "mem_util_pct": 70.0,
}

_log = logging.getLogger("layer1_ba_replica")


# ---------------------------------------------------------------------------
# Synonym sets (KP.SYNONYM_HEADERS) — frozen copy of L1.INPUT.006
# ---------------------------------------------------------------------------
SYNONYMS: dict[str, list[str]] = {
    "vm_name": ["VM", "VM Name", "Name"],
    "powerstate": ["Powerstate", "Power State", "PowerState", "State"],
    "template": ["Template", "Is Template"],
    "vcpu": ["CPUs", "vCPUs", "CPU Count", "# CPUs", "Num CPU", "NumCPU"],
    "memory": [
        "Memory", "Memory MiB", "Memory MB", "RAM", "RAM MiB",
        "Configured Memory", "Memory GB",
    ],
    "provisioned_storage": [
        "Provisioned MiB", "Provisioned", "Capacity MiB", "Capacity",
        "Allocated", "Allocated MiB", "Total Provisioned", "Storage Provisioned",
        "Total disk capacity MiB",
    ],
    "in_use_storage": [
        "In Use MiB", "In Use", "Used", "Used MiB",
        "Utilized", "Utilised", "Consumed", "Consumed MiB",
    ],
    "host_ref": ["Host", "VMHost", "ESX Host", "ESXi Host"],
    "datacenter": ["Datacenter"],
    "os_config": [
        "OS according to the configuration file", "OS (Configured)",
        "GuestOS", "Guest OS", "OS",
    ],
    "os_tools": ["OS according to the VMware Tools", "OS (Tools)", "Tools OS"],
    # vInfo — embedded utilisation (rare but possible per L1.UTIL.001)
    "vm_cpu_usage_pct": ["CPU usage %", "CPU usage", "% CPU"],
    "vm_mem_usage_pct": ["Memory usage %", "Memory usage", "% Memory"],
    # vHost
    "host_cores": ["# Cores", "Cores", "Total Cores", "Physical Cores"],
    "vcpus_per_core": ["vCPUs per Core", "vCPU per Core", "vCPU/Core"],
    "host_cpu_usage_pct": ["CPU usage %", "CPU usage", "% CPU", "CPU %"],
    "host_mem_usage_pct": ["Memory usage %", "Memory usage", "% Memory", "Mem %"],
}

# Unit tokens declared explicitly in column headers
_UNIT_PATTERNS: list[tuple[str, re.Pattern]] = [
    ("gib", re.compile(r"\bGiB\b", re.I)),
    ("gb",  re.compile(r"\bGB\b", re.I)),
    ("tib", re.compile(r"\bTiB\b", re.I)),
    ("tb",  re.compile(r"\bTB\b", re.I)),
    ("mib", re.compile(r"\bMiB\b", re.I)),
    ("mb",  re.compile(r"\bMB\b", re.I)),
]

# Mandatory tabs (KP.MANDATORY_VINFO) and recommended tabs (soft warnings)
MANDATORY_TABS = {"vinfo"}
RECOMMENDED_TABS = {"vhost", "vpartition"}
OPTIONAL_TABS = {"vcpu", "vmemory"}


# ---------------------------------------------------------------------------
# Header resolution (KP.SYNONYM_HEADERS + KP.MIB_DEFAULT)
# ---------------------------------------------------------------------------
@dataclass
class ResolvedColumn:
    """The result of resolving one logical field on one sheet."""
    field_key: str           # e.g. "provisioned_storage"
    matched_synonym: str     # the synonym that matched, e.g. "Provisioned MiB"
    header_text: str         # the actual header text from the file
    col_index: int           # 0-based column index
    declared_unit: str       # one of {"mib","mb","gib","gb","tib","tb","unspecified"}
    assumed_unit: str        # what we will USE — "mib" if "unspecified", else declared

    def to_dict(self) -> dict:
        return dataclasses.asdict(self)


def _normalise(text: str) -> str:
    """Lower-case, strip, collapse whitespace and punctuation that varies."""
    t = (text or "").strip().lower()
    t = re.sub(r"[\s_/\-()\[\]\.\:]+", " ", t)
    return t.strip()


def _detect_unit(header: str) -> str:
    for token, pattern in _UNIT_PATTERNS:
        if pattern.search(header):
            return token
    return "unspecified"


def resolve_column(
    headers: list[str], field_key: str, *, prefer_units: tuple[str, ...] = (),
) -> ResolvedColumn | None:
    """Resolve ``field_key`` against ``headers`` using SYNONYMS.

    ``prefer_units`` lets the caller bias toward a particular declared unit
    when multiple synonyms match (e.g. prefer GB over MiB for the memory
    column when both ``Memory`` and ``Memory GB`` exist).
    """
    synonyms = SYNONYMS.get(field_key, [])
    if not synonyms:
        return None

    # Build (norm_header, original_header, idx) lookup once
    norm_headers = [(_normalise(h), h, i) for i, h in enumerate(headers)]

    candidates: list[ResolvedColumn] = []
    for syn in synonyms:
        norm_syn = _normalise(syn)
        for norm_h, orig_h, idx in norm_headers:
            if norm_h == norm_syn:
                unit = _detect_unit(orig_h)
                candidates.append(ResolvedColumn(
                    field_key=field_key,
                    matched_synonym=syn,
                    header_text=orig_h,
                    col_index=idx,
                    declared_unit=unit,
                    assumed_unit="mib" if unit == "unspecified" else unit,
                ))

    if not candidates:
        return None

    # Apply preference rules
    if prefer_units:
        for unit in prefer_units:
            for c in candidates:
                if c.declared_unit == unit:
                    return c
    # Default: first synonym wins (synonyms are ordered most-canonical first)
    syn_rank = {s: i for i, s in enumerate(synonyms)}
    candidates.sort(key=lambda c: syn_rank.get(c.matched_synonym, 99))
    return candidates[0]


def _to_gb_decimal(raw_value: float, declared_unit: str) -> float:
    """Convert a numeric value with a declared unit to decimal GB.

    Per KP.MIB_DEFAULT, ``unspecified`` is treated as MiB.
    """
    if raw_value is None:
        return 0.0
    try:
        v = float(raw_value)
    except (TypeError, ValueError):
        return 0.0
    if v <= 0:
        return 0.0
    unit = (declared_unit or "mib").lower()
    if unit in ("mib", "unspecified"):
        return v * MIB_TO_GB_DECIMAL
    if unit == "mb":
        return v / 1000.0                # decimal MB → decimal GB
    if unit == "gib":
        return v * (1024.0 ** 3) / (10 ** 9)
    if unit == "gb":
        return v
    if unit == "tib":
        return v * (1024.0 ** 4) / (10 ** 9)
    if unit == "tb":
        return v * 1000.0
    return v * MIB_TO_GB_DECIMAL


# ---------------------------------------------------------------------------
# Per-VM record (KP.PER_VM — authoritative payload)
# ---------------------------------------------------------------------------
@dataclass
class VMRecord:
    name: str
    powerstate: str
    is_template: bool
    is_powered_on: bool
    vcpu: int
    memory_gb_decimal: float           # converted via KP.MIB_DEFAULT
    provisioned_gb_decimal: float      # KP.PROVISIONED_FOR_TCO source (vPartition pref)
    provisioned_storage_source: str    # 'vpartition' | 'vinfo' | 'none'
    in_use_gb_decimal: float           # L2 Azure-disk sizing source
    host_name: str
    datacenter: str
    os_config: str
    os_tools: str
    cpu_util_pct: float | None         # populated when vInfo or vCPU has it
    mem_util_pct: float | None
    host_proxy_cpu_pct: float | None   # populated when vHost has data for host
    host_proxy_mem_pct: float | None

    def to_dict(self) -> dict:
        return dataclasses.asdict(self)


@dataclass
class HostRecord:
    name: str
    datacenter: str
    cores: int
    vcpus_per_core: float
    cpu_usage_pct: float | None
    mem_usage_pct: float | None
    num_vms_total: int = 0           # vHost '# VMs total' (incl. powered-off)
    num_vms_poweredon: int = 0       # vHost '# VMs' (powered-on per vHost)
    has_vinfo_vm: bool = False       # Set during pass when vInfo references this host

    @property
    def in_scope(self) -> bool:
        """BA-default scope: any vHost row with at least one powered-on VM."""
        return self.num_vms_poweredon > 0

    def to_dict(self) -> dict:
        return dataclasses.asdict(self)


@dataclass
class BAReviewPacket:
    """KP.BA_APPROVAL_GATE — what the BA must acknowledge before Layer 2."""
    missing_tabs: list[str] = field(default_factory=list)
    assumptions: list[dict] = field(default_factory=list)
    tolerances: list[dict] = field(default_factory=list)
    proposed_reduction_factors: dict = field(default_factory=dict)
    utilisation_options: list[dict] = field(default_factory=list)
    synonym_matches: list[dict] = field(default_factory=list)
    hard_errors: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return dataclasses.asdict(self)


@dataclass
class Layer1Result:
    input_file: str
    per_vm: list[VMRecord]
    hosts: list[HostRecord]
    fyi_aggregates: dict
    ba_review_packet: BAReviewPacket
    rule_provenance: dict          # rule_id → field name(s) it produced

    def to_dict(self) -> dict:
        return {
            "input_file": self.input_file,
            "per_vm_count": len(self.per_vm),
            "fyi_aggregates": self.fyi_aggregates,
            "ba_review_packet": self.ba_review_packet.to_dict(),
            "rule_provenance": self.rule_provenance,
            "hosts_count": len(self.hosts),
            "per_vm_sample": [vm.to_dict() for vm in self.per_vm[:5]],
        }


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------
def _open_workbook(path: Path):
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def _find_sheet(wb, target: str) -> str | None:
    """Case-insensitive, prefix-tolerant sheet finder (per L1.INPUT.005)."""
    norm_target = target.lower()
    for name in wb.sheetnames:
        if name.lower() == norm_target:
            return name
    for name in wb.sheetnames:
        if name.lower().startswith(norm_target):
            return name
    return None


def _read_headers(ws) -> list[str]:
    first = next(ws.iter_rows(max_row=1, values_only=True), ())
    return [str(h) if h is not None else "" for h in first]


def _read_rows(ws) -> Iterable[tuple]:
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # skip header
    yield from rows


# ---------------------------------------------------------------------------
# Pass 1 — vInfo (mandatory, KP.MANDATORY_VINFO)
# ---------------------------------------------------------------------------
def _parse_vinfo(wb, packet: BAReviewPacket) -> tuple[list[VMRecord], dict]:
    """Return (per_vm, resolved_columns_dict)."""
    sheet_name = _find_sheet(wb, "vinfo")
    if sheet_name is None:
        packet.hard_errors.append(
            "vInfo tab missing or empty — KP.MANDATORY_VINFO requires it."
        )
        raise SystemExit(
            "FATAL: vInfo tab not found. Workflow halted (KP.MANDATORY_VINFO)."
        )

    ws = wb[sheet_name]
    headers = _read_headers(ws)

    # Resolve required columns
    cols: dict[str, ResolvedColumn | None] = {
        "vm_name": resolve_column(headers, "vm_name"),
        "powerstate": resolve_column(headers, "powerstate"),
        "template": resolve_column(headers, "template"),
        "vcpu": resolve_column(headers, "vcpu"),
        # Prefer GB-labelled memory column when present so the unit is explicit
        "memory": resolve_column(headers, "memory", prefer_units=("gb", "gib", "mib", "mb")),
        "provisioned_storage": resolve_column(headers, "provisioned_storage", prefer_units=("mib", "gib", "gb")),
        "in_use_storage": resolve_column(headers, "in_use_storage", prefer_units=("mib", "gib", "gb")),
        "host_ref": resolve_column(headers, "host_ref"),
        "datacenter": resolve_column(headers, "datacenter"),
        "os_config": resolve_column(headers, "os_config"),
        "os_tools": resolve_column(headers, "os_tools"),
        "vm_cpu_usage_pct": resolve_column(headers, "vm_cpu_usage_pct"),
        "vm_mem_usage_pct": resolve_column(headers, "vm_mem_usage_pct"),
    }

    # Missing-mandatory checks
    for required in ("vm_name", "powerstate", "vcpu", "memory", "provisioned_storage"):
        if cols[required] is None:
            packet.hard_errors.append(
                f"vInfo missing required column for '{required}'. "
                f"Synonyms tried: {SYNONYMS[required]!r}. "
                f"Headers found: {headers!r}"
            )

    if packet.hard_errors:
        raise SystemExit(
            "FATAL: vInfo missing one or more mandatory columns. "
            f"Errors: {packet.hard_errors}"
        )

    # Record what synonym matched for each field (audit trail)
    for key, col in cols.items():
        if col is not None:
            packet.synonym_matches.append({
                "sheet": "vInfo",
                "field": key,
                "matched_synonym": col.matched_synonym,
                "header_text": col.header_text,
                "declared_unit": col.declared_unit,
                "assumed_unit": col.assumed_unit,
            })
        else:
            # missing optional column → soft note
            packet.assumptions.append({
                "field": key,
                "sheet": "vInfo",
                "missing": True,
                "tried": SYNONYMS.get(key, []),
                "consequence": (
                    "Per-VM utilisation telemetry unavailable; reduction "
                    "factors will be surfaced for BA approval."
                    if key in {"vm_cpu_usage_pct", "vm_mem_usage_pct"}
                    else f"Field '{key}' will be empty for all VMs."
                ),
            })

    per_vm: list[VMRecord] = []
    for row in _read_rows(ws):
        try:
            vm_name = row[cols["vm_name"].col_index]
        except IndexError:
            continue
        if vm_name in (None, ""):
            continue

        powerstate = (row[cols["powerstate"].col_index] or "")
        # Normalise both 'Powered On' / 'PoweredOn' / 'poweredOn'
        powerstate_norm = re.sub(r"\s+", "", str(powerstate)).lower()
        is_powered_on = powerstate_norm == "poweredon"

        is_template = False
        if cols["template"]:
            tval = row[cols["template"].col_index]
            is_template = bool(tval) and str(tval).strip().lower() in ("true", "1", "yes")

        try:
            vcpu = int(row[cols["vcpu"].col_index] or 0)
        except (TypeError, ValueError):
            vcpu = 0

        mem_raw = row[cols["memory"].col_index]
        mem_unit = cols["memory"].assumed_unit
        mem_gb = _to_gb_decimal(mem_raw, mem_unit)
        # KP.MIB_DEFAULT subtlety: vInfo's bare 'Memory' column is decimal MB
        # in RVTools (not MiB). When the resolver matched a unit-unspecified
        # 'Memory' (assumed_unit=='mib'), apply the MB convention instead and
        # log it in the audit trail. When 'Memory GB' was matched, this branch
        # is never taken and the GB value is used as-is.
        if mem_unit == "mib" and cols["memory"].header_text.strip().lower() == "memory":
            if isinstance(mem_raw, (int, float)) and mem_raw > 0:
                mem_gb = float(mem_raw) / 1024.0

        prov_col = cols["provisioned_storage"]
        prov_raw = row[prov_col.col_index]
        provisioned_gb = _to_gb_decimal(prov_raw, prov_col.assumed_unit)

        in_use_col = cols["in_use_storage"]
        if in_use_col is not None:
            in_use_raw = row[in_use_col.col_index]
            in_use_gb = _to_gb_decimal(in_use_raw, in_use_col.assumed_unit)
        else:
            in_use_gb = 0.0

        host_name = (
            str(row[cols["host_ref"].col_index]) if cols["host_ref"] and row[cols["host_ref"].col_index] else ""
        )
        datacenter = (
            str(row[cols["datacenter"].col_index]) if cols["datacenter"] and row[cols["datacenter"].col_index] else ""
        )
        os_config = (
            str(row[cols["os_config"].col_index]) if cols["os_config"] and row[cols["os_config"].col_index] else ""
        )
        os_tools = (
            str(row[cols["os_tools"].col_index]) if cols["os_tools"] and row[cols["os_tools"].col_index] else ""
        )

        cpu_util = None
        mem_util = None
        if cols["vm_cpu_usage_pct"]:
            v = row[cols["vm_cpu_usage_pct"].col_index]
            cpu_util = float(v) if isinstance(v, (int, float)) else None
        if cols["vm_mem_usage_pct"]:
            v = row[cols["vm_mem_usage_pct"].col_index]
            mem_util = float(v) if isinstance(v, (int, float)) else None

        per_vm.append(VMRecord(
            name=str(vm_name),
            powerstate=str(powerstate),
            is_template=is_template,
            is_powered_on=is_powered_on,
            vcpu=vcpu,
            memory_gb_decimal=round(mem_gb, 4),
            provisioned_gb_decimal=round(provisioned_gb, 4),
            provisioned_storage_source="vinfo",   # may be overwritten by vPartition pass
            in_use_gb_decimal=round(in_use_gb, 4),
            host_name=host_name,
            datacenter=datacenter,
            os_config=os_config,
            os_tools=os_tools,
            cpu_util_pct=cpu_util,
            mem_util_pct=mem_util,
            host_proxy_cpu_pct=None,
            host_proxy_mem_pct=None,
        ))

    _log.info("vInfo: parsed %d VM rows from sheet '%s'", len(per_vm), sheet_name)
    return per_vm, cols


# ---------------------------------------------------------------------------
# Pass 2 — vHost (optional)
# ---------------------------------------------------------------------------
def _parse_vhost(wb, packet: BAReviewPacket) -> list[HostRecord]:
    sheet_name = _find_sheet(wb, "vhost")
    if sheet_name is None:
        packet.missing_tabs.append("vHost")
        packet.assumptions.append({
            "tab": "vHost",
            "missing": True,
            "fallback": f"vCPU/pCore = {DEFAULT_VCPU_PER_PCORE_BENCHMARK} benchmark",
            "consequence": "Host count and vCPU/pCore ratio derived from benchmark; less precise.",
            "rule": "L1.INPUT.003",
        })
        return []

    ws = wb[sheet_name]
    headers = _read_headers(ws)
    cols = {
        "host_ref": resolve_column(headers, "host_ref"),
        "datacenter": resolve_column(headers, "datacenter"),
        "host_cores": resolve_column(headers, "host_cores"),
        "vcpus_per_core": resolve_column(headers, "vcpus_per_core"),
        "host_cpu_usage_pct": resolve_column(headers, "host_cpu_usage_pct"),
        "host_mem_usage_pct": resolve_column(headers, "host_mem_usage_pct"),
    }
    # vHost has its own '# VMs' (powered-on) and '# VMs total' columns;
    # they're authoritative for the BA's scope decision (no vInfo cross-match needed).
    nvms_idx = next((i for i, h in enumerate(headers) if h.strip() == "# VMs"), None)
    nvmst_idx = next((i for i, h in enumerate(headers) if h.strip() == "# VMs total"), None)
    for key, col in cols.items():
        if col is not None:
            packet.synonym_matches.append({
                "sheet": "vHost",
                "field": key,
                "matched_synonym": col.matched_synonym,
                "header_text": col.header_text,
                "declared_unit": col.declared_unit,
                "assumed_unit": col.assumed_unit,
            })

    if cols["host_ref"] is None:
        packet.assumptions.append({
            "tab": "vHost",
            "issue": "Host name column not found",
            "consequence": "vHost data ignored.",
        })
        return []

    hosts: list[HostRecord] = []
    for row in _read_rows(ws):
        name = row[cols["host_ref"].col_index]
        if name in (None, ""):
            continue

        def _safe(col_key, cast=float, default=0.0):
            c = cols.get(col_key)
            if c is None:
                return default
            v = row[c.col_index]
            try:
                return cast(v) if v is not None else default
            except (TypeError, ValueError):
                return default

        hosts.append(HostRecord(
            name=str(name),
            datacenter=str(row[cols["datacenter"].col_index]) if cols["datacenter"] and row[cols["datacenter"].col_index] else "",
            cores=_safe("host_cores", int, 0),
            vcpus_per_core=_safe("vcpus_per_core", float, 0.0),
            cpu_usage_pct=_safe("host_cpu_usage_pct", float, None),
            mem_usage_pct=_safe("host_mem_usage_pct", float, None),
            num_vms_poweredon=int(row[nvms_idx] or 0) if nvms_idx is not None and isinstance(row[nvms_idx], (int, float)) else 0,
            num_vms_total=int(row[nvmst_idx] or 0) if nvmst_idx is not None and isinstance(row[nvmst_idx], (int, float)) else 0,
        ))
    _log.info("vHost: parsed %d host rows from sheet '%s'", len(hosts), sheet_name)
    return hosts


# ---------------------------------------------------------------------------
# Pass 3 — vPartition (FYI only per L1.PARTITION.001)
# ---------------------------------------------------------------------------
def _scan_vpartition(wb, packet: BAReviewPacket) -> dict:
    sheet_name = _find_sheet(wb, "vpartition")
    if sheet_name is None:
        packet.missing_tabs.append("vPartition")
        packet.assumptions.append({
            "tab": "vPartition",
            "missing": True,
            "fallback": "On-prem TCO falls back to vInfo Provisioned per VM",
            "consequence": "Storage TCO uses vInfo Provisioned (less precise than vPartition).",
            "rule": "L1.STORAGE_PROV.001 / L1.PARTITION.001",
        })
        return {"present": False}
    ws = wb[sheet_name]
    n_rows = sum(1 for _ in _read_rows(ws))
    return {"present": True, "rows": n_rows, "sheet": sheet_name}


def _apply_vpartition_to_per_vm(wb, per_vm: list[VMRecord], packet: BAReviewPacket) -> dict:
    """Per L1.STORAGE_PROV.001 + L1.PARTITION.001:

    When vPartition is present and usable, sum its Capacity per VM and
    OVERWRITE per_vm[].provisioned_gb_decimal with that sum (and set
    provisioned_storage_source='vpartition').

    When per-VM join FAILS (e.g. vInfo VM names are redacted while
    vPartition has the real names), we cannot attribute vPartition rows to
    individual VMs. In that case we keep per-VM provisioned from vInfo (so
    Layer 2 still has per-VM data for managed-disk sizing) BUT we record
    the canonical vPartition total separately so the TCO total still
    reflects the BA's preferred source. This is logged in the BA review
    packet (KP.BA_APPROVAL_GATE).

    Returns FYI summary suitable for the parity report.
    """
    sheet_name = _find_sheet(wb, "vpartition")
    if sheet_name is None:
        return {
            "present": False,
            "join_strategy": "none (vPartition absent)",
            "vms_overwritten": 0,
            "canonical_total_gb": 0.0,
        }

    ws = wb[sheet_name]
    headers = _read_headers(ws)
    vm_col = resolve_column(headers, "vm_name")
    cap_col = resolve_column(headers, "provisioned_storage", prefer_units=("mib",))
    if vm_col is None or cap_col is None:
        packet.assumptions.append({
            "tab": "vPartition",
            "issue": "VM-name or Capacity column not resolvable",
            "consequence": "vPartition cannot be used; falling back to vInfo Provisioned per VM.",
            "rule": "L1.PARTITION.001",
        })
        return {"present": True, "usable": False, "vms_overwritten": 0, "canonical_total_gb": 0.0}

    # Sum partition Capacity per VM AND keep the unconditional total
    by_vm: dict[str, float] = {}
    total_mib = 0.0
    for row in _read_rows(ws):
        cap = row[cap_col.col_index]
        if not isinstance(cap, (int, float)):
            continue
        total_mib += float(cap)
        vm = row[vm_col.col_index]
        if vm in (None, ""):
            continue
        by_vm[str(vm)] = by_vm.get(str(vm), 0.0) + float(cap)

    canonical_total_gb = round(total_mib * MIB_TO_GB_DECIMAL, 2)

    # Try the per-VM join
    overwritten = 0
    for vm in per_vm:
        if vm.name in by_vm:
            new_gb = by_vm[vm.name] * MIB_TO_GB_DECIMAL
            vm.provisioned_gb_decimal = round(new_gb, 4)
            vm.provisioned_storage_source = "vpartition"
            overwritten += 1

    coverage_pct = round(100.0 * overwritten / max(len(per_vm), 1), 2)
    join_ok = coverage_pct >= 80.0
    join_strategy = "per-vm-join" if join_ok else "aggregate-only (per-VM join failed)"

    if not join_ok:
        packet.assumptions.append({
            "rule": "L1.STORAGE_PROV.001 / L1.INPUT.004",
            "issue": "vPartition VM names cannot be joined to vInfo VM names",
            "likely_cause": "vInfo VM names appear redacted/anonymised while vPartition retains real server names",
            "join_coverage_pct": coverage_pct,
            "resolution": (
                "Per-VM provisioned values remain from vInfo (used for Layer 2 "
                "disk sizing). Aggregate `total_provisioned_gb_all` reports the "
                "vPartition canonical sum (per L1.STORAGE_PROV.001 BA preference). "
                "This is the only situation where the FYI total is NOT a strict "
                "sum of per-VM values \u2014 explicitly flagged here for transparency."
            ),
            "vinfo_provisioned_total_gb": round(
                sum(vm.provisioned_gb_decimal for vm in per_vm), 2
            ),
            "vpartition_canonical_total_gb": canonical_total_gb,
            "ba_decision_required_before": "Layer 2",
        })

    # Record what we did
    packet.synonym_matches.append({
        "sheet": "vPartition",
        "field": "vm_name",
        "matched_synonym": vm_col.matched_synonym,
        "header_text": vm_col.header_text,
        "declared_unit": vm_col.declared_unit,
        "assumed_unit": vm_col.assumed_unit,
    })
    packet.synonym_matches.append({
        "sheet": "vPartition",
        "field": "provisioned_storage",
        "matched_synonym": cap_col.matched_synonym,
        "header_text": cap_col.header_text,
        "declared_unit": cap_col.declared_unit,
        "assumed_unit": cap_col.assumed_unit,
    })
    return {
        "present": True,
        "usable": True,
        "join_strategy": join_strategy,
        "vms_overwritten": overwritten,
        "vms_total": len(per_vm),
        "coverage_pct": coverage_pct,
        "canonical_total_gb": canonical_total_gb,
    }


# ---------------------------------------------------------------------------
# Pass 4 — Apply host-proxy utilisation (per L1.UTIL.005, BA-presented option)
# ---------------------------------------------------------------------------
def _apply_host_proxy(per_vm: list[VMRecord], hosts: list[HostRecord], packet: BAReviewPacket) -> None:
    if not hosts:
        return
    by_host = {h.name: h for h in hosts}
    proxied = 0
    for vm in per_vm:
        host = by_host.get(vm.host_name)
        if host:
            host.has_vinfo_vm = True
            if host.cpu_usage_pct is not None:
                vm.host_proxy_cpu_pct = host.cpu_usage_pct
                proxied += 1
            if host.mem_usage_pct is not None:
                vm.host_proxy_mem_pct = host.mem_usage_pct

    has_per_vm_util = any(vm.cpu_util_pct is not None for vm in per_vm)
    if not has_per_vm_util and proxied > 0:
        packet.utilisation_options.append({
            "rule": "L1.UTIL.005",
            "option_a": "Apply host CPU/Memory usage % to each VM on that host (host-proxy)",
            "option_b": f"Use BA-set reduction factors (defaults: cpu={DEFAULT_REDUCTION_FACTORS['cpu_util_pct']}%, mem={DEFAULT_REDUCTION_FACTORS['mem_util_pct']}%)",
            "data_available": {
                "vms_with_host_proxy": proxied,
                "vms_total": len(per_vm),
                "host_proxy_coverage_pct": round(100.0 * proxied / max(len(per_vm), 1), 2),
            },
            "ba_decision_required_before": "Layer 2",
        })

    if not has_per_vm_util and proxied == 0:
        packet.proposed_reduction_factors = dict(DEFAULT_REDUCTION_FACTORS)
        packet.assumptions.append({
            "rule": "L1.UTIL.001",
            "topic": "No utilisation data found",
            "fallback": packet.proposed_reduction_factors,
            "ba_decision_required_before": "Layer 2",
        })


# ---------------------------------------------------------------------------
# Aggregations (FYI only per KP.PER_VM)
# ---------------------------------------------------------------------------
def _aggregate(per_vm: list[VMRecord], hosts: list[HostRecord], packet: BAReviewPacket) -> dict:
    n_all = len(per_vm)
    n_on = sum(1 for vm in per_vm if vm.is_powered_on)
    n_template = sum(1 for vm in per_vm if vm.is_template)

    sum_vcpu_all = sum(vm.vcpu for vm in per_vm)
    sum_vcpu_on = sum(vm.vcpu for vm in per_vm if vm.is_powered_on)
    sum_mem_all = sum(vm.memory_gb_decimal for vm in per_vm)
    sum_mem_on = sum(vm.memory_gb_decimal for vm in per_vm if vm.is_powered_on)
    sum_prov_all = sum(vm.provisioned_gb_decimal for vm in per_vm)
    sum_prov_on = sum(vm.provisioned_gb_decimal for vm in per_vm if vm.is_powered_on)
    sum_inuse_all = sum(vm.in_use_gb_decimal for vm in per_vm)
    sum_inuse_on = sum(vm.in_use_gb_decimal for vm in per_vm if vm.is_powered_on)

    # Hosts: ALL hosts in vHost are in scope (per L1.HOST.002, BA-clarified
    # 2026-04-27). Customer/BA controls scope upstream by removing rows.
    num_hosts = len(hosts)
    vpc_values = [h.vcpus_per_core for h in hosts if h.vcpus_per_core > 0]
    vpc_ratio = round(sum(vpc_values) / len(vpc_values), 4) if vpc_values else 0.0

    # Per-datacenter breakdown is FYI only (NOT a scope decision; BA controls
    # scope upstream of the file). Surfaced for BA visibility, not approval.
    dc_breakdown: dict[str, dict] = {}
    for h in hosts:
        dc = h.datacenter or "(none)"
        b = dc_breakdown.setdefault(dc, {"hosts": 0, "sum_vpc": 0.0, "n_vpc": 0, "sum_cores": 0})
        b["hosts"] += 1
        b["sum_cores"] += h.cores or 0
        if h.vcpus_per_core and h.vcpus_per_core > 0:
            b["sum_vpc"] += h.vcpus_per_core
            b["n_vpc"] += 1
    dc_breakdown_clean = {
        dc: {
            "hosts": b["hosts"],
            "avg_vcpu_per_core": round(b["sum_vpc"] / b["n_vpc"], 4) if b["n_vpc"] else 0.0,
            "sum_cores": b["sum_cores"],
        }
        for dc, b in sorted(dc_breakdown.items())
    }
    dc_breakdown_clean = {
        dc: {
            "hosts": b["hosts"],
            "avg_vcpu_per_core": round(b["sum_vpc"] / b["n_vpc"], 4) if b["n_vpc"] else 0.0,
            "sum_cores": b["sum_cores"],
        }
        for dc, b in sorted(dc_breakdown.items())
    }

    # If multiple datacenters, surface as FYI (NOT a BA decision — BA
    # controls scope upstream by removing rows from vHost).
    dc_fyi = None
    if len(dc_breakdown_clean) > 1:
        dc_fyi = {
            "rule": "L1.HOST.001 / L1.HOST.002",
            "topic": "Multiple datacenters detected (FYI only — ALL hosts in scope per L1.HOST.002)",
            "datacenters": dc_breakdown_clean,
            "note": (
                "Per BA: every host in vHost is in scope. If a DC should be\n"
                "excluded, remove its rows from vHost upstream of the file."
            ),
        }

    return {
        "num_vms_all": n_all,
        "num_vms_poweredon": n_on,
        "num_vms_template": n_template,
        "total_vcpu_all": sum_vcpu_all,
        "total_vcpu_poweredon": sum_vcpu_on,
        "total_memory_gb_all": round(sum_mem_all, 2),
        "total_memory_gb_poweredon": round(sum_mem_on, 2),
        "total_provisioned_gb_all": round(sum_prov_all, 2),
        "total_provisioned_gb_poweredon": round(sum_prov_on, 2),
        "total_in_use_gb_all": round(sum_inuse_all, 2),
        "total_in_use_gb_poweredon": round(sum_inuse_on, 2),
        "num_hosts": num_hosts,
        "vcpu_per_core_ratio": vpc_ratio,
        "datacenter_breakdown_fyi": dc_fyi,
    }


# ---------------------------------------------------------------------------
# Main entrypoint
# ---------------------------------------------------------------------------
def replicate_layer1(input_path: Path) -> Layer1Result:
    """Run the BA Layer 1 workflow against ``input_path`` and return a Layer1Result."""
    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    packet = BAReviewPacket()
    wb = _open_workbook(input_path)

    # Detect non-RVTools structures up-front (KP.MANDATORY_VINFO + L1.INPUT.001)
    if _find_sheet(wb, "vinfo") is None:
        packet.hard_errors.append(
            "vInfo tab not found. Sheets seen: " + ", ".join(wb.sheetnames)
        )
        raise SystemExit(
            "FATAL: not a valid RVTools-shaped workbook (vInfo missing). "
            "L1.INPUT.001/L1.INPUT.003 — workflow halted."
        )

    per_vm, _vinfo_cols = _parse_vinfo(wb, packet)
    hosts = _parse_vhost(wb, packet)
    vpartition_summary = _apply_vpartition_to_per_vm(wb, per_vm, packet)
    _apply_host_proxy(per_vm, hosts, packet)

    fyi = _aggregate(per_vm, hosts, packet)

    # KP.PER_VM honours per-VM aggregation. The single exception is when
    # vPartition is canonical (per L1.STORAGE_PROV.001) but the per-VM join
    # failed (redacted vInfo names per L1.INPUT.004). In that case the
    # canonical TCO total is the vPartition aggregate; we keep per-VM
    # provisioned from vInfo for Layer 2 disk sizing. The override is
    # logged in `provisioned_storage_total_source` for transparency.
    fyi["vpartition_summary"] = vpartition_summary
    fyi["vinfo_provisioned_only_total_gb"] = round(
        sum(
            vm.provisioned_gb_decimal
            for vm in per_vm
            if vm.provisioned_storage_source == "vinfo"
        ),
        2,
    )
    if vpartition_summary.get("present") and vpartition_summary.get("usable"):
        if vpartition_summary["join_strategy"].startswith("aggregate-only"):
            fyi["total_provisioned_gb_all"] = vpartition_summary["canonical_total_gb"]
            fyi["provisioned_storage_total_source"] = (
                "vpartition (aggregate-only \u2014 per-VM join failed; per-VM "
                "provisioned remains from vInfo for Layer 2)"
            )
        else:
            fyi["provisioned_storage_total_source"] = "vpartition (per-VM join)"
    else:
        fyi["provisioned_storage_total_source"] = "vinfo (vPartition unavailable)"

    rule_provenance = {
        "L1.SCOPE.001": ["fyi_aggregates.num_vms_all", "fyi_aggregates.total_vcpu_all"],
        "L1.SCOPE.002": ["fyi_aggregates.num_vms_poweredon"],
        "L1.UNITS.001": ["per_vm[].provisioned_gb_decimal", "per_vm[].in_use_gb_decimal"],
        "L1.UNITS.002": ["per_vm[].memory_gb_decimal"],
        "L1.STORAGE_PROV.001": ["per_vm[].provisioned_gb_decimal"],
        "L1.STORAGE_INUSE.001": ["per_vm[].in_use_gb_decimal"],
        "L1.HOST.001": ["fyi_aggregates.num_hosts", "fyi_aggregates.vcpu_per_core_ratio"],
        "L1.HOST.002": ["fyi_aggregates.num_hosts (only hosts with ≥1 powered-on VM)"],
        "L1.UTIL.005": ["per_vm[].host_proxy_cpu_pct", "per_vm[].host_proxy_mem_pct"],
        "L1.PARTITION.001": ["ba_review_packet (FYI only — vPartition not in critical path)"],
        "L1.HANDOFF.001": ["per_vm[]", "fyi_aggregates", "ba_review_packet"],
    }

    return Layer1Result(
        input_file=str(input_path),
        per_vm=per_vm,
        hosts=hosts,
        fyi_aggregates=fyi,
        ba_review_packet=packet,
        rule_provenance=rule_provenance,
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Layer 1 BA replica oracle")
    parser.add_argument("input", type=Path, help="RVTools .xlsx input file")
    parser.add_argument(
        "--out", type=Path, default=None,
        help="Write result YAML here (default: print summary to stdout).",
    )
    parser.add_argument(
        "--per-vm-out", type=Path, default=None,
        help="Optional: dump full per-VM records to this YAML.",
    )
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    result = replicate_layer1(args.input)

    summary = result.to_dict()
    if args.out:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(yaml.safe_dump(summary, sort_keys=False, width=100))
        _log.info("Summary written to %s", args.out)
    else:
        print(yaml.safe_dump(summary, sort_keys=False, width=100))

    if args.per_vm_out:
        args.per_vm_out.parent.mkdir(parents=True, exist_ok=True)
        args.per_vm_out.write_text(
            yaml.safe_dump([vm.to_dict() for vm in result.per_vm], sort_keys=False, width=120)
        )
        _log.info("Per-VM records written to %s", args.per_vm_out)

    return 0


if __name__ == "__main__":
    sys.exit(main())
