"""
Layer 3 Golden Extractor
========================

Extracts every authoritative Business Analyst output cell from a finalised
``BV Benchmark Business Case`` workbook into a typed Python oracle.

This module is the **single source of truth** for Layer 3 parity testing:

    BA spreadsheet (oracle, frozen)
            │
            ▼
    ┌───────────────────────┐     ┌───────────────────────┐
    │   Layer-3 Replica     │ ←─► │   engine/financial_*  │
    │ (formula-faithful Py) │     │ (production code)     │
    └───────────────────────┘     └───────────────────────┘

Both replica and engine must match the oracle to within the tolerance bands
defined in ``layer3_judge.py``. The judge runs a **3-way audit** so any
divergence (replica↔BA, engine↔BA, or replica↔engine) is surfaced.

Every extracted value carries a ``CellRef`` (sheet, cell address, label) so
the auditor can point back at the exact cell in the BA's workbook when a
discrepancy is found — there is never any ambiguity about provenance.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Sequence

import openpyxl
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Provenance: every golden value carries its origin coordinate
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class CellRef:
    """A single cell coordinate in the BA workbook."""

    sheet: str
    address: str  # e.g. "C10"
    label: str  # human-readable description

    def __str__(self) -> str:  # pragma: no cover - display only
        return f"{self.sheet}!{self.address} ({self.label})"


@dataclass(frozen=True)
class GoldenValue:
    """A scalar BA-authoritative value with provenance."""

    ref: CellRef
    value: float


@dataclass(frozen=True)
class GoldenSeries:
    """A Y0..Y10 series with one cell per year."""

    label: str
    sheet: str
    row: int
    col_y0: int  # 1-based column index for Y0
    values: tuple[float, ...]  # length 11

    def cell_for_year(self, year_index: int) -> CellRef:
        col_letter = get_column_letter(self.col_y0 + year_index)
        return CellRef(self.sheet, f"{col_letter}{self.row}", f"{self.label} Y{year_index}")


# ---------------------------------------------------------------------------
# Top-level oracle structure
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class StatusQuoGolden:
    """Status Quo P&L line items (Detailed Financial Case rows 8-37)."""

    server_depreciation: GoldenSeries
    server_hw_maintenance: GoldenSeries
    storage_depreciation: GoldenSeries
    storage_maintenance: GoldenSeries
    storage_backup: GoldenSeries
    storage_dr: GoldenSeries
    nw_fitout_depreciation: GoldenSeries
    network_hw_maintenance: GoldenSeries
    bandwidth_costs: GoldenSeries
    dc_lease_space: GoldenSeries
    dc_power: GoldenSeries
    virtualization_licenses: GoldenSeries
    windows_licenses: GoldenSeries
    sql_licenses: GoldenSeries
    windows_esu: GoldenSeries
    sql_esu: GoldenSeries
    backup_licenses: GoldenSeries
    dr_licenses: GoldenSeries
    it_admin_staff: GoldenSeries
    total_on_prem_cost: GoldenSeries  # row 37 — sum across all SQ rows


@dataclass(frozen=True)
class CashFlowGolden:
    """Cash-flow view (Summary Financial Case rows 14-29)."""

    sq_capex: GoldenSeries
    sq_opex: GoldenSeries
    sq_total: GoldenSeries
    az_capex: GoldenSeries
    az_opex: GoldenSeries
    az_consumption: GoldenSeries
    az_migration: GoldenSeries
    az_ms_funding: GoldenSeries
    az_total: GoldenSeries
    savings: GoldenSeries
    cf_delta: GoldenSeries  # az - sq (negative means saving)
    cf_rate: GoldenSeries


@dataclass(frozen=True)
class HeadlineGolden:
    """Top-line summary metrics (Summary Financial Case rows 6-12)."""

    npv_sq_10y: GoldenValue
    npv_sq_5y: GoldenValue
    npv_az_10y: GoldenValue
    npv_az_5y: GoldenValue
    terminal_value_10y: GoldenValue
    terminal_value_5y: GoldenValue
    project_npv_with_tv_10y: GoldenValue
    project_npv_with_tv_5y: GoldenValue
    project_npv_excl_tv_10y: GoldenValue
    project_npv_excl_tv_5y: GoldenValue
    roi_5y_cf: GoldenValue
    payback_years: GoldenValue
    y10_savings_10y_cf: GoldenValue
    y10_savings_5y_cf: GoldenValue
    y10_savings_rate_10y: GoldenValue
    y10_savings_rate_5y: GoldenValue


@dataclass(frozen=True)
class FivePaybackGolden:
    """5Y CF with Payback sheet (the displayed ROI/Payback method)."""

    infra_cost_reduction_npv: GoldenValue  # H8
    infra_admin_reduction_npv: GoldenValue  # H10
    total_benefits_npv: GoldenValue  # H18
    incremental_azure_npv: GoldenValue  # H22
    migration_npv: GoldenValue  # H24
    total_costs_npv: GoldenValue  # H28
    net_benefits_npv: GoldenValue  # H30 (= Summary D10)
    roi_5y_cf: GoldenValue  # H31 (= Summary E6)
    payback_years: GoldenValue  # I32 / E11


@dataclass(frozen=True)
class DetailedNPVGolden:
    """Detailed Financial Case NPV section (rows 91-101)."""

    annual_npv_y1: GoldenValue
    annual_npv_y10: GoldenValue
    npv_10y_excl_tv: GoldenValue  # M94
    terminal_value_10y_raw: GoldenValue  # N93 (gross)
    npv_with_tv_10y_raw: GoldenValue  # N98
    wacc: GoldenValue
    perpetual_growth_rate: GoldenValue


@dataclass(frozen=True)
class StatusQuoEstimationGolden:
    """Status Quo Estimation tab — acquisition costs and DC details."""

    server_acquisition_cost: GoldenValue  # B7
    storage_acquisition_cost: GoldenValue  # B19
    nw_fitout_acquisition_cost: GoldenValue  # B32
    licenses_yearly_cost: GoldenValue  # B40
    dc_lease_space_yearly: GoldenValue  # B59
    dc_power_yearly: GoldenValue  # B60
    bandwidth_yearly: GoldenValue  # B86
    server_hw_maint_yearly: GoldenValue  # B94
    storage_hw_maint_yearly: GoldenValue  # B102
    network_hw_maint_yearly: GoldenValue  # B110
    sysadmin_yearly: GoldenValue  # baseline


@dataclass(frozen=True)
class LayerThreeGolden:
    """Complete Layer 3 oracle for a single customer."""

    customer_name: str
    workbook_path: str

    status_quo: StatusQuoGolden
    cash_flow: CashFlowGolden
    headline: HeadlineGolden
    five_payback: FivePaybackGolden
    detailed_npv: DetailedNPVGolden
    sq_estimation: StatusQuoEstimationGolden

    # Flat dict for ad-hoc auditing / debugging
    raw_cells: dict[str, GoldenValue] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Extractor
# ---------------------------------------------------------------------------


def _read_series(
    ws,
    label: str,
    row: int,
    col_y0: int = 3,  # column C (1-based) — Y0 in Summary Financial Case
    n_years: int = 11,
) -> GoldenSeries:
    """Read 11 consecutive cells (Y0..Y10) from one row."""
    values = []
    for i in range(n_years):
        v = ws.cell(row=row, column=col_y0 + i).value
        values.append(float(v) if isinstance(v, (int, float)) else 0.0)
    return GoldenSeries(label=label, sheet=ws.title, row=row, col_y0=col_y0, values=tuple(values))


def _read_cell(ws, address: str, label: str) -> GoldenValue:
    cell = ws[address]
    val = cell.value
    if not isinstance(val, (int, float)):
        # Strings like "Proportional" can't be cast — caller must handle separately
        val = 0.0
    return GoldenValue(ref=CellRef(ws.title, address, label), value=float(val))


def extract_layer3_golden(workbook_path: str | Path) -> LayerThreeGolden:
    """Pull every BA-authoritative Layer 3 cell from ``workbook_path``."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True, keep_vba=False)

    sf = wb["Summary Financial Case"]  # column C = Y0, M = Y10, N = 10Y total
    df = wb["Detailed Financial Case"]  # column B = Y0, L = Y10 — different layout!
    fp = wb["5Y CF with Payback"]
    sq = wb["Status Quo Estimation"]

    # Customer name (A1 of Summary)
    customer = (sf["B1"].value or "").split(" Cloud")[0].strip() or "Unknown"

    # ------------------------------------------------------------------
    # Status Quo P&L line items — Detailed Financial Case rows 8..37
    # In this sheet: column B = label, C = Y0 baseline, D = Y1, ... M = Y10
    # so col_y0 = 3 (column C). N = 10y total (separate scalar).
    # ------------------------------------------------------------------
    status_quo = StatusQuoGolden(
        server_depreciation=_read_series(df, "Server Depreciation", row=8, col_y0=3),
        server_hw_maintenance=_read_series(df, "Server HW Maintenance", row=9, col_y0=3),
        storage_depreciation=_read_series(df, "Storage Depreciation", row=10, col_y0=3),
        storage_maintenance=_read_series(df, "Storage Maintenance", row=11, col_y0=3),
        storage_backup=_read_series(df, "Storage Backup", row=12, col_y0=3),
        storage_dr=_read_series(df, "Storage DR", row=13, col_y0=3),
        nw_fitout_depreciation=_read_series(df, "NW+Fitout Depreciation", row=14, col_y0=3),
        network_hw_maintenance=_read_series(df, "Network HW Maintenance", row=15, col_y0=3),
        bandwidth_costs=_read_series(df, "Bandwidth Costs", row=18, col_y0=3),
        dc_lease_space=_read_series(df, "DC Lease (Space)", row=20, col_y0=3),
        dc_power=_read_series(df, "DC Power", row=21, col_y0=3),
        virtualization_licenses=_read_series(df, "Virtualization Licenses", row=24, col_y0=3),
        windows_licenses=_read_series(df, "Windows Server Licenses", row=25, col_y0=3),
        sql_licenses=_read_series(df, "SQL Server Licenses", row=26, col_y0=3),
        windows_esu=_read_series(df, "Windows Server ESU", row=27, col_y0=3),
        sql_esu=_read_series(df, "SQL Server ESU", row=28, col_y0=3),
        backup_licenses=_read_series(df, "Backup Licenses", row=29, col_y0=3),
        dr_licenses=_read_series(df, "Disaster Recovery Licenses", row=30, col_y0=3),
        it_admin_staff=_read_series(df, "IT Admin Staff", row=33, col_y0=3),
        total_on_prem_cost=_read_series(df, "Total On-Prem Cost", row=37, col_y0=3),
    )

    # ------------------------------------------------------------------
    # Cash flow view — Summary Financial Case rows 14..29
    # In this sheet: column C = Y0, M = Y10 (col 3..13)
    # ------------------------------------------------------------------
    cash_flow = CashFlowGolden(
        sq_capex=_read_series(sf, "SQ CAPEX", row=16, col_y0=3),
        sq_opex=_read_series(sf, "SQ OPEX", row=17, col_y0=3),
        sq_total=_read_series(sf, "SQ Total CF", row=19, col_y0=3),
        az_capex=_read_series(sf, "AZ CAPEX", row=21, col_y0=3),
        az_opex=_read_series(sf, "AZ OPEX", row=22, col_y0=3),
        az_consumption=_read_series(sf, "AZ Consumption", row=23, col_y0=3),
        az_migration=_read_series(sf, "AZ Migration", row=24, col_y0=3),
        az_ms_funding=_read_series(sf, "AZ MS Funding", row=25, col_y0=3),
        az_total=_read_series(sf, "AZ Total CF", row=26, col_y0=3),
        savings=_read_series(sf, "Savings (SQ-AZ)", row=27, col_y0=3),
        cf_delta=_read_series(sf, "CF Delta (AZ-SQ)", row=28, col_y0=3),
        cf_rate=_read_series(sf, "CF Rate", row=29, col_y0=3),
    )

    # ------------------------------------------------------------------
    # Headline metrics — Summary rows 6..12
    # ------------------------------------------------------------------
    headline = HeadlineGolden(
        npv_sq_10y=_read_cell(sf, "C6", "NPV Status Quo 10y"),
        npv_sq_5y=_read_cell(sf, "D6", "NPV Status Quo 5y"),
        npv_az_10y=_read_cell(sf, "C7", "NPV Azure Case 10y"),
        npv_az_5y=_read_cell(sf, "D7", "NPV Azure Case 5y"),
        terminal_value_10y=_read_cell(sf, "C8", "Terminal Value 10y"),
        terminal_value_5y=_read_cell(sf, "D8", "Terminal Value 5y"),
        project_npv_with_tv_10y=_read_cell(sf, "C9", "Project NPV w/ TV 10y"),
        project_npv_with_tv_5y=_read_cell(sf, "D9", "Project NPV w/ TV 5y"),
        project_npv_excl_tv_10y=_read_cell(sf, "C10", "Project NPV excl TV 10y"),
        project_npv_excl_tv_5y=_read_cell(sf, "D10", "Project NPV excl TV 5y"),
        roi_5y_cf=_read_cell(sf, "E6", "ROI 5Y CF"),
        payback_years=_read_cell(sf, "E11", "Payback Years"),
        y10_savings_10y_cf=_read_cell(sf, "C11", "Y10 Savings (10y CF)"),
        y10_savings_5y_cf=_read_cell(sf, "D11", "Y10 Savings (5y CF)"),
        y10_savings_rate_10y=_read_cell(sf, "C12", "Y10 Savings Rate (10y CF)"),
        y10_savings_rate_5y=_read_cell(sf, "D12", "Y10 Savings Rate (5y CF)"),
    )

    # ------------------------------------------------------------------
    # 5Y CF with Payback sheet
    # Column H = undiscounted Total, Column I = discounted (NPV at WACC).
    # The headline "Net Benefits NPV" lives in I30; the headline "ROI" in I31.
    # ------------------------------------------------------------------
    five_payback = FivePaybackGolden(
        infra_cost_reduction_npv=_read_cell(fp, "H8", "Infra Cost Reduction NPV"),
        infra_admin_reduction_npv=_read_cell(fp, "H10", "Infra Admin Reduction NPV"),
        total_benefits_npv=_read_cell(fp, "H18", "Total Benefits NPV"),
        incremental_azure_npv=_read_cell(fp, "H22", "Incremental Azure NPV"),
        migration_npv=_read_cell(fp, "H24", "Migration NPV"),
        total_costs_npv=_read_cell(fp, "H28", "Total Costs NPV"),
        net_benefits_npv=_read_cell(fp, "I30", "Net Benefits NPV"),
        roi_5y_cf=_read_cell(fp, "I31", "ROI 5Y CF"),
        payback_years=_read_cell(fp, "I32", "Payback Years"),
    )

    # ------------------------------------------------------------------
    # Detailed NPV section (rows 91-101)
    # In this sheet, NPV row 94 cols D..M = annual discounted Y1..Y10,
    # N94 = 10Y total NPV excl TV, O94 = 5Y total NPV excl TV.
    # ------------------------------------------------------------------
    detailed_npv = DetailedNPVGolden(
        annual_npv_y1=_read_cell(df, "D94", "Annual NPV Y1"),
        annual_npv_y10=_read_cell(df, "M94", "Annual NPV Y10"),
        npv_10y_excl_tv=_read_cell(df, "N94", "NPV 10y excl TV"),
        terminal_value_10y_raw=_read_cell(df, "M93", "Terminal Value 10y (gross)"),
        npv_with_tv_10y_raw=_read_cell(df, "N98", "NPV w/ TV 10y total"),
        wacc=_read_cell(df, "C100", "WACC"),
        perpetual_growth_rate=_read_cell(df, "C101", "Perpetual Growth Rate"),
    )

    # ------------------------------------------------------------------
    # Status Quo Estimation tab — acquisition + yearly baseline costs.
    # Column B = label, C = total/yearly value, D = depr life, E = depr cost.
    # Sysadmin baseline (Y0): I166 = total cost; I148 = per-FTE rate.
    # ------------------------------------------------------------------
    sq_est = StatusQuoEstimationGolden(
        server_acquisition_cost=_read_cell(sq, "C7", "Server Acquisition Cost"),
        storage_acquisition_cost=_read_cell(sq, "C19", "Storage Acquisition Cost"),
        nw_fitout_acquisition_cost=_read_cell(sq, "C32", "NW+Fitout Acquisition Cost"),
        licenses_yearly_cost=_read_cell(sq, "C40", "Licenses Yearly Cost"),
        dc_lease_space_yearly=_read_cell(sq, "C59", "DC Lease Space Yearly"),
        dc_power_yearly=_read_cell(sq, "C60", "DC Power Yearly"),
        bandwidth_yearly=_read_cell(sq, "C86", "Bandwidth Yearly"),
        server_hw_maint_yearly=_read_cell(sq, "C94", "Server HW Maint Yearly"),
        storage_hw_maint_yearly=_read_cell(sq, "C102", "Storage HW Maint Yearly"),
        network_hw_maint_yearly=_read_cell(sq, "C110", "Network HW Maint Yearly"),
        sysadmin_yearly=_read_cell(sq, "I166", "Sysadmin Total Y0 Cost"),
    )

    return LayerThreeGolden(
        customer_name=customer,
        workbook_path=str(workbook_path),
        status_quo=status_quo,
        cash_flow=cash_flow,
        headline=headline,
        five_payback=five_payback,
        detailed_npv=detailed_npv,
        sq_estimation=sq_est,
    )


# ---------------------------------------------------------------------------
# Convenience: flatten golden into (label, ref, value) tuples
# ---------------------------------------------------------------------------


def flatten_golden(g: LayerThreeGolden) -> list[tuple[str, CellRef, float]]:
    """
    Flatten the entire oracle into ``(label, cell_ref, value)`` tuples.

    Used by ``layer3_judge.py`` to compare against replica/engine outputs
    cell-by-cell.
    """
    out: list[tuple[str, CellRef, float]] = []

    # Series — emit one tuple per (series, year)
    series_blocks: Sequence[tuple[str, Iterable[GoldenSeries]]] = (
        ("status_quo", _series_iter(g.status_quo)),
        ("cash_flow", _series_iter(g.cash_flow)),
    )
    for prefix, series_iter in series_blocks:
        for s in series_iter:
            for yr, val in enumerate(s.values):
                out.append((f"{prefix}.{s.label}.Y{yr}", s.cell_for_year(yr), val))

    # Scalars
    for prefix, obj in (
        ("headline", g.headline),
        ("five_payback", g.five_payback),
        ("detailed_npv", g.detailed_npv),
        ("sq_estimation", g.sq_estimation),
    ):
        for fname in obj.__dataclass_fields__:
            gv: GoldenValue = getattr(obj, fname)
            out.append((f"{prefix}.{fname}", gv.ref, gv.value))

    return out


def _series_iter(block) -> Iterable[GoldenSeries]:
    for fname in block.__dataclass_fields__:
        yield getattr(block, fname)
