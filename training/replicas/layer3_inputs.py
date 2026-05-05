"""
Layer 3 BA Replica — Customer-Workbook Inputs Loader
=====================================================

Reads the BA's hand-typed inputs from the ``1-Client Variables`` sheet of
a finalised business-case workbook into a typed Python dataclass tree.
The replica modules consume this dataclass; they never reach back into
the workbook for inputs themselves.

Inputs are deliberately split between:

* ``InputsClient``        — what the BA enters per engagement
* ``InputsBenchmark``     — values from ``Benchmark Assumptions`` (constants)
* ``InputsConsumption``   — Azure run-rate plan (Layer 3.5)

Everything is plain Python primitives (float / str / bool); no Pydantic
or other heavy machinery so the replica stays drift-free from the engine.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class InputsClient:
    """Per-engagement inputs from ``1-Client Variables``."""

    # Engagement
    client_name: str
    currency: str

    # Pricing & Datacenter
    win_price_level: str  # "B" or "D"
    sql_price_level: str
    nb_dc: int
    dc_exit_type: str  # "Static" or "Proportional"
    nb_interconnects: int

    # Hardware lifecycle
    hw_depr_life_yrs: float  # default 5
    hw_actual_life_yrs: float  # default 5
    expected_future_growth_rate: float  # e.g. 0.02
    hw_renewal_during_migration_pct: float  # e.g. 0.10
    incorporate_productivity: bool

    # Workload #1 — physical totals
    workload_name: str
    nb_vms: float
    nb_physical_servers: float
    total_vms_and_servers_combined: float  # D41 = VMs + physical servers (excl. VM hosts) — used for migration cost
    allocated_vcpu: float
    allocated_pcores: float
    allocated_vmem_gb: float
    allocated_pmem_gb: float
    allocated_storage_gb: float

    # Backup / DR
    backup_size_gb: float
    backup_protected_vms: float
    dr_size_gb: float
    dr_protected_vms: float

    # License pCore counts
    pcores_with_virtualization: float
    byol_virtualization: bool
    vcpu_per_pcore_ratio: float
    pcores_with_windows_server: float
    pcores_with_windows_server_esu: float
    pcores_with_sql_server: float
    pcores_with_sql_server_esu: float


@dataclass(frozen=True)
class InputsBenchmark:
    """Benchmark constants (column K of Benchmark Assumptions)."""

    wacc: float
    hours_per_year: float
    watt_to_kwh: float
    gb_to_tb: float

    # Servers & storage
    vm_to_physical_server_ratio: float
    vmem_to_pmem_ratio: float
    server_cost_per_core: float
    server_cost_per_gb_memory: float
    storage_cost_per_gb: float
    gb_storage_already_in_servers: float
    backup_storage_cost_per_gb_yr: float
    dr_storage_cost_per_gb_yr: float
    server_hw_maintenance_pct: float
    storage_hw_maintenance_pct: float

    # Network & fitout
    servers_per_cabinet: float
    core_routers_per_dc: float
    aggregate_routers_per_core: float
    access_switches_per_core: float
    load_balancers_per_core: float
    cabinet_cost: float
    core_router_cost: float
    aggregate_router_cost: float
    access_switch_cost: float
    load_balancer_cost: float
    network_hw_maintenance_pct: float

    # Licenses
    windows_server_license_per_core_yr_b: float
    sql_server_license_per_core_yr_b: float
    windows_esu_per_core_yr_b: float
    sql_esu_per_core_yr_b: float
    windows_server_license_per_core_yr_d: float
    sql_server_license_per_core_yr_d: float
    windows_esu_per_core_yr_d: float
    sql_esu_per_core_yr_d: float
    virtualization_license_per_core_yr: float
    backup_software_per_vm_yr: float
    dr_software_per_vm_yr: float

    # Lease & power
    unused_power_overhead_pct: float
    space_cost_per_kw_month: float
    power_cost_per_kw_month: float
    on_prem_pue: float
    tdp_watt_per_core: float
    storage_power_kwh_yr_per_tb: float
    on_prem_load_factor: float

    # Bandwidth & telecom
    interconnect_cost_per_yr: float

    # IT admin
    vms_per_sysadmin: float
    sysadmin_fully_loaded_cost_yr: float
    sysadmin_working_hours_yr: float
    sysadmin_contractor_pct: float

    # Productivity
    productivity_reduction_after_migration: float
    productivity_recapture_rate: float

    # Terminal value
    perpetual_growth_rate: float

    # NII
    nii_interest_rate: float = 0.03


@dataclass(frozen=True)
class InputsConsumption:
    """Per-workload Azure consumption plan (only Workload #1 fully populated for Customer A)."""

    migration_cost_per_vm: float  # default 1500
    migration_ramp_eoy: tuple[float, ...]  # length 10 (Y1..Y10), values 0..1

    # Y10 anchor consumption (annual run rate, list-cost)
    compute_consumption_y10: float
    storage_consumption_y10: float
    other_consumption_y10: float

    # Discount + funding
    azure_consumption_discount: float  # ACD, 0..1
    aco_total: float  # one-time, derived from per-year sums (informational)
    ecif_total: float  # one-time, derived from per-year sums (informational)
    # Per-year ACO/ECIF (Step 15 / Customer B): authoritative source for funding.
    # BA workbook canonical: '2a-Consumption Plan Wk1'!E21:N21 (ACO) and E22:N22 (ECIF).
    # D21 / D22 are SUM() formulas over these per-year cells. For Customer A all
    # per-year cells are blank (treated as 0.0); for Customer B, Y1..Y3 carry the
    # ECIF subsidy. Length is exactly 10 = Y1..Y10.
    aco_by_year: tuple[float, ...]
    ecif_by_year: tuple[float, ...]

    # Backup / DR
    backup_activated: bool
    backup_storage_in_azure_run_rate: bool
    backup_software_in_azure_run_rate: bool
    dr_activated: bool
    dr_storage_in_azure_run_rate: bool
    dr_software_in_azure_run_rate: bool


# ---------------------------------------------------------------------------
# Loader
# ---------------------------------------------------------------------------


def _f(cell, default: float = 0.0) -> float:
    v = cell.value
    return float(v) if isinstance(v, (int, float)) else default


def _s(cell, default: str = "") -> str:
    v = cell.value
    return str(v).strip() if v is not None else default


def _b(cell) -> bool:
    v = cell.value
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().lower() in ("yes", "true", "y", "1")
    return False


def load_client_inputs(workbook_path: str | Path) -> InputsClient:
    """Read ``1-Client Variables`` into ``InputsClient``."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True, keep_vba=False)
    cv = wb["1-Client Variables"]

    return InputsClient(
        client_name=_s(cv["D9"], "Customer"),
        currency=_s(cv["D10"], "USD"),
        win_price_level=_s(cv["D13"], "B"),
        sql_price_level=_s(cv["D14"], "B"),
        nb_dc=int(_f(cv["D18"])),
        dc_exit_type=_s(cv["D19"], "Proportional"),
        nb_interconnects=int(_f(cv["D20"])),
        hw_depr_life_yrs=_f(cv["D24"], 5.0),
        hw_actual_life_yrs=_f(cv["D25"], 5.0),
        expected_future_growth_rate=_f(cv["D26"], 0.02),
        hw_renewal_during_migration_pct=_f(cv["D27"], 0.10),
        incorporate_productivity=_b(cv["D31"]),
        workload_name=_s(cv["D35"], "Workload #1"),
        nb_vms=_f(cv["D39"]),
        # D40 is "excl. VM hosts" (often 0); D42 is "incl. VM hosts" (the BA-canonical number)
        nb_physical_servers=_f(cv["D42"]),
        # D41 is the BA's "Total VM's and physical servers combined" — used directly in migration formula.
        total_vms_and_servers_combined=_f(cv["D41"]),
        allocated_vcpu=_f(cv["D44"]),
        allocated_pcores=_f(cv["D47"]),
        allocated_vmem_gb=_f(cv["D49"]),
        allocated_pmem_gb=_f(cv["D52"]),
        allocated_storage_gb=_f(cv["D54"]),
        backup_size_gb=_f(cv["D58"]),
        backup_protected_vms=_f(cv["D59"]),
        dr_size_gb=_f(cv["D60"]),
        dr_protected_vms=_f(cv["D61"]),
        pcores_with_virtualization=_f(cv["D64"]),
        byol_virtualization=_b(cv["D65"]),
        vcpu_per_pcore_ratio=_f(cv["D66"], 1.97),
        pcores_with_windows_server=_f(cv["D67"]),
        pcores_with_windows_server_esu=_f(cv["D68"]),
        pcores_with_sql_server=_f(cv["D70"]),
        pcores_with_sql_server_esu=_f(cv["D71"]),
    )


def load_benchmark_inputs(workbook_path: str | Path) -> InputsBenchmark:
    """Read column K of ``Benchmark Assumptions`` into ``InputsBenchmark``."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True, keep_vba=False)
    ba = wb["Benchmark Assumptions"]
    K = lambda r: _f(ba.cell(row=r, column=11))  # noqa: E731 - column K = "Value Used"

    return InputsBenchmark(
        wacc=K(5),
        hours_per_year=K(6),
        watt_to_kwh=K(7),
        gb_to_tb=K(8),
        vm_to_physical_server_ratio=K(11),
        vmem_to_pmem_ratio=K(13),
        server_cost_per_core=K(14),
        server_cost_per_gb_memory=K(15),
        storage_cost_per_gb=K(16),
        gb_storage_already_in_servers=K(17),
        backup_storage_cost_per_gb_yr=K(18),
        dr_storage_cost_per_gb_yr=K(19),
        server_hw_maintenance_pct=K(20),
        storage_hw_maintenance_pct=K(21),
        servers_per_cabinet=K(24),
        core_routers_per_dc=K(25),
        aggregate_routers_per_core=K(26),
        access_switches_per_core=K(27),
        load_balancers_per_core=K(28),
        cabinet_cost=K(30),
        core_router_cost=K(31),
        aggregate_router_cost=K(32),
        access_switch_cost=K(33),
        load_balancer_cost=K(34),
        network_hw_maintenance_pct=K(35),
        windows_server_license_per_core_yr_b=K(38),
        sql_server_license_per_core_yr_b=K(39),
        windows_esu_per_core_yr_b=K(40),
        sql_esu_per_core_yr_b=K(41),
        windows_server_license_per_core_yr_d=K(42),
        sql_server_license_per_core_yr_d=K(43),
        windows_esu_per_core_yr_d=K(44),
        sql_esu_per_core_yr_d=K(45),
        virtualization_license_per_core_yr=K(46),
        backup_software_per_vm_yr=K(47),
        dr_software_per_vm_yr=K(48),
        unused_power_overhead_pct=K(51),
        space_cost_per_kw_month=K(52),
        power_cost_per_kw_month=K(53),
        on_prem_pue=K(54),
        tdp_watt_per_core=K(55),
        storage_power_kwh_yr_per_tb=K(56),
        on_prem_load_factor=K(57),
        interconnect_cost_per_yr=K(60),
        vms_per_sysadmin=K(63),
        sysadmin_fully_loaded_cost_yr=K(64),
        sysadmin_working_hours_yr=K(65),
        sysadmin_contractor_pct=K(66),
        productivity_reduction_after_migration=K(67),
        productivity_recapture_rate=K(68),
        # Perpetual growth & NII rate live on Detailed sheet — passed via override
        perpetual_growth_rate=0.03,
        nii_interest_rate=0.03,
    )


def load_consumption_inputs(workbook_path: str | Path) -> InputsConsumption:
    """Read ``2a-Consumption Plan Wk1`` into ``InputsConsumption``."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True, keep_vba=False)
    cp = wb["2a-Consumption Plan Wk1"]

    # Migration ramp Y1..Y10 — row 17, cols E..N (5..14)
    ramp = tuple(_f(cp.cell(row=17, column=c)) for c in range(5, 15))

    # Per-year ACO (row 21) and ECIF (row 22), cols E..N (5..14). Blank → 0.0.
    # BA D21 = SUM(E21:N21) and D22 = SUM(E22:N22) — per-year is canonical.
    aco_by_year = tuple(_f(cp.cell(row=21, column=c)) for c in range(5, 15))
    ecif_by_year = tuple(_f(cp.cell(row=22, column=c)) for c in range(5, 15))

    return InputsConsumption(
        migration_cost_per_vm=_f(cp["D14"], 1500.0),
        migration_ramp_eoy=ramp,
        compute_consumption_y10=_f(cp["N28"]),
        storage_consumption_y10=_f(cp["N29"]),
        other_consumption_y10=_f(cp["N30"]),
        azure_consumption_discount=_f(cp["D8"]),  # if present
        aco_total=_f(cp["D21"]),
        ecif_total=_f(cp["D22"]),
        aco_by_year=aco_by_year,
        ecif_by_year=ecif_by_year,
        backup_activated=_b(cp["E35"]),
        backup_storage_in_azure_run_rate=_b(cp["E38"]),
        backup_software_in_azure_run_rate=_b(cp["E39"]),
        dr_activated=_b(cp["E42"]),
        dr_storage_in_azure_run_rate=_b(cp["E45"]),
        dr_software_in_azure_run_rate=_b(cp["E46"]),
    )
