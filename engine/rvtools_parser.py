"""
RVtools export parser.

Reads a standard RVTools .xlsx export and extracts the fields needed
to populate a WorkloadInventory for the business case engine.

RVTools column names are used (not positional indices) so the parser
is robust across different RVTools export versions.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Column name constants
# ---------------------------------------------------------------------------

import logging
_log = logging.getLogger(__name__)

# vInfo columns
COL_VM_NAME = "VM"
COL_POWERSTATE = "Powerstate"
COL_TEMPLATE = "Template"
COL_CPUS = "CPUs"
COL_MEMORY_MB = "Memory"          # megabytes
COL_IN_USE_MIB = "In Use MiB"     # storage in use
COL_OS_CONFIG = "OS according to the configuration file"
COL_OS_TOOLS = "OS according to the VMware Tools"

# vHost columns
COL_HOST = "Host"
COL_CORES = "# Cores"            # total physical cores per host
COL_HOST_MEMORY_MB = "# Memory"  # megabytes
COL_VCPUS_PER_CORE = "vCPUs per Core"

# OS filter patterns for Windows Server
_WINDOWS_PATTERN = re.compile(r"windows\s+server", re.IGNORECASE)

# Environment tags that explicitly indicate non-production workloads.
# Any SQL/Windows VM whose Environment column contains one of these
# keywords is classified as non-production.  All other VMs — including
# those with an empty, unknown, or "production" environment tag — are
# classified as production.  This implements the default assumption:
# "if no tagging, assume production."
_ENV_NONPROD_PATTERN = re.compile(
    r"\b(dev|development|test|testing|uat|qa|staging|sandbox|non.?prod)\b",
    re.IGNORECASE,
)

# ESU-eligible OS versions: 2003, 2008 (inc R2), and 2012 (inc R2).
# Windows Server 2012/2012 R2 reached end of standard support Oct 2023 and
# became ESU-eligible.  2003/2008/2008 R2 have been ESU for longer.
# NOTE: RVtools' 'OS according to the configuration file' column often shows
# a generic string (e.g. 'Microsoft Windows Server 2016 or later') for older
# VMs whose VMware hardware version predates current OS detection.  The VMware
# Tools column is used as a fallback but may also be unreliable for retired VMs.
# When unversioned Windows Server VMs are detected, a warning is emitted.
_WINDOWS_ESU_PATTERN = re.compile(
    r"windows\s+server\s+(2003|2008|2012)", re.IGNORECASE
)

# Any Windows Server OS string that includes a 4-digit year (e.g. 2016, 2019, 2022)
# is considered "versioned".  VMs without any year in either OS column are
# "unversioned" — VMware could not determine the exact version — and may be
# older than 2016, so ESU exposure may be understated.
_WINDOWS_VERSIONED_PATTERN = re.compile(
    r"windows\s+server\s+\d{4}", re.IGNORECASE
)

MIB_TO_GB = 1 / 953.67   # MiB → decimal GB  (for on-prem TCO storage totals only)
MIB_TO_GIB = 1 / 1024.0  # MiB → GiB  (binary; matches Azure SKU catalog memory/disk units)
MB_TO_GB = 1 / 1024.0


# ---------------------------------------------------------------------------
# Per-VM record (populated during vInfo pass; used for per-VM rightsizing)
# ---------------------------------------------------------------------------

@dataclass
class VMRecord:
    """
    Lightweight per-VM record extracted from vInfo (+ disk sizes from vDisk).
    Only powered-on, non-template VMs are included.
    All memory/storage values are in GiB (MiB ÷ 1024) to match Azure catalog.
    """
    name: str
    vcpu: int
    memory_mib: int         # raw MiB from vInfo 'Memory' column
    host_name: str          # vInfo 'Host' column (for vHost proxy lookup)
    os_cfg: str
    os_tools: str
    app_str: str            # 'Application' custom attribute (lowercased)
    is_windows: bool
    is_esu: bool
    is_sql: bool
    # Per-disk provisioned sizes in GiB (from vDisk 'Capacity MiB' ÷ 1024).
    # Populated in a second pass after vDisk is parsed; defaults to empty.
    disk_sizes_gib: list[float] = field(default_factory=list)
    # vPartition consumed GiB (sum of all partition 'Consumed MiB' ÷ 1024 for this VM)
    # 0.0 means vPartition tab absent or no match.
    partition_consumed_gib: float = 0.0
    # vInfo fallback storage values (GiB), used when vDisk/vPartition absent
    inuse_gib: float = 0.0
    provisioned_gib: float = 0.0


# ---------------------------------------------------------------------------
# Raw parsed result
# ---------------------------------------------------------------------------

@dataclass
class RVToolsInventory:
    """
    Aggregated inventory derived from a single RVTools export.

    Two counting scopes:

      ON-PREM TCO BASELINE — all VMs (powered-on + powered-off), excl. templates:
        num_vms, total_vcpu, total_vmemory_gb, total_storage_in_use_gb,
        pcores_with_windows_server, pcores_with_windows_esu.
        Rationale: the customer paid for all hardware and software licences
        regardless of whether individual VMs are currently running.

      AZURE MIGRATION TARGET — powered-on VMs only:
        num_vms_poweredon, total_vcpu_poweredon, total_vmemory_gb_poweredon,
        total_storage_poweredon_gb.
        Rationale: only running VMs are candidates for migration; powered-off
        VMs will not consume Azure resources.

      ESU detection: both OS columns ('configuration file' and 'VMware Tools')
        are checked.  The configuration-file column is the primary source as it
        more frequently carries explicit version strings.  When Windows Server
        VMs with no detectable version are found (typically pre-2016 VMs whose
        VMware hardware version predates OS detection), esu_count_may_be_understated
        is set to True and a warning is emitted.
    """
    # ── ON-PREM TCO BASELINE (all VMs, incl. powered-off, excl. templates) ──
    num_vms: int = 0
    total_vcpu: int = 0
    total_vmemory_gb: float = 0.0
    total_storage_in_use_gb: float = 0.0  # all VMs, In Use MiB / 953.67

    # Host / virtualisation layer (from vHost tab)
    vcpu_per_core_ratio: float = 0.0
    total_host_pcores: int = 0
    total_host_memory_gb: float = 0.0
    num_hosts: int = 0

    # Windows/SQL pCore estimates — ALL VMs, both OS columns
    pcores_with_windows_server: int = 0
    pcores_with_windows_esu: int = 0
    esu_count_may_be_understated: bool = False
    windows_vms_unknown_version: int = 0  # Windows VMs with no detectable version
    pcores_with_sql_server: int = 0       # default: 10% of windows; overridden by Application detection
    pcores_with_sql_esu: int = 0          # default: 10% of windows_esu

    # SQL detection from Application custom attribute (col 77 in vInfo).
    # When sql_vms_detected > 0 it overrides the 10% Windows default above.
    # Prod/Non-Prod split is informational (for presentation); not used in TCO math.
    #
    # Default assumption: if a VM has no environment tag (or an unrecognised tag),
    # it is classified as Production.  Only explicit non-production tags
    # (dev/test/uat/staging/sandbox/non-prod) override this to Non-Production.
    sql_vms_detected: int = 0             # VMs with 'sql' in Application or OS
    sql_vms_prod: int = 0                 # Production (explicit tag or assumed)
    sql_vms_nonprod: int = 0              # Non-Production (explicit non-prod tag only)
    sql_prod_assumed: bool = False        # True when no SQL VM had any env tag
    sql_detection_source: str = "default"  # 'application' | 'default'

    # Environment tagging coverage
    env_tagging_present: bool = False     # True if ≥1 VM has any non-empty Environment tag

    # ── AZURE MIGRATION TARGET (powered-on VMs only) ──
    num_vms_poweredon: int = 0
    total_vcpu_poweredon: int = 0
    total_vmemory_gb_poweredon: float = 0.0
    total_storage_poweredon_gb: float = 0.0  # powered-on In Use MiB / 953.67 (on-prem TCO basis)

    # Provisioned disk capacity from vDisk tab — used for Azure managed disk cost estimation.
    # Azure bills on provisioned tier size, not consumed bytes.  If vDisk tab is absent
    # these remain 0.0 and consumption_builder falls back to in-use × headroom.
    total_disk_provisioned_gb: float = 0.0        # all VMs, vDisk.Capacity MiB sum / 953.67
    total_disk_provisioned_poweredon_gb: float = 0.0  # powered-on VMs only
    # Per-VM disk layout for per-VM managed disk tier costing.
    # Keys are VM names; values are lists of provisioned sizes in GiB (float).
    # Populated only for powered-on, non-template VMs.  Empty when vDisk tab absent.
    vm_disk_sizes_gb: dict[str, list[float]] = field(default_factory=dict)

    # ── UTILISATION TELEMETRY (from vCPU and vMemory tabs) ──
    # Values are fleet P95 fractions (0–1+); 0.0 = telemetry not available.
    # powered-on VMs only; powered-off VMs (Overall==0) are excluded.
    cpu_util_p95: float = 0.0           # P95 of (Overall MHz / Max MHz) per VM
    cpu_util_p95_vm_count: int = 0      # VMs contributing to the P95 calculation
    memory_util_p95: float = 0.0        # P95 of (Consumed MiB / Size MiB) per VM
    memory_util_p95_vm_count: int = 0

    # ── REGION EVIDENCE (from vHost and vMetaData tabs) ──
    # These raw signals are consumed by engine/region_guesser.py to infer an
    # Azure region.  All collections are de-duplicated and sorted.
    datacenter_names: list[str] = field(default_factory=list)         # vHost.Datacenter (unique, sorted)
    datacenter_host_counts: dict[str, int] = field(default_factory=dict)  # {dc_name: host_count}
    timezone_names: list[str] = field(default_factory=list)           # vHost.Time Zone Name
    gmt_offsets: list[str] = field(default_factory=list)              # vHost.GMT Offset (as strings)
    domain_names: list[str] = field(default_factory=list)             # vHost.Domain
    vcenter_fqdns: list[str] = field(default_factory=list)            # vMetaData.Server

    # Parse context
    vhost_available: bool = False           # True when vHost tab was present
    include_powered_off_applied: bool = False  # True = all-VMs TCO baseline used

    # Per-VM records for rightsizing (powered-on, non-template VMs only)
    # Populated during parse(); empty until vInfo is processed.
    vm_records: list["VMRecord"] = field(default_factory=list)

    # Per-VM utilisation maps (from vCPU / vMemory tabs)
    # Keys are VM names; values are utilisation fractions (0–1+).
    # 0.0 entries are not stored — absence means no telemetry.
    vm_cpu_util: dict[str, float] = field(default_factory=dict)
    vm_mem_util: dict[str, float] = field(default_factory=dict)

    # VM → Host mapping (from vInfo 'Host' column)
    vm_to_host: dict[str, str] = field(default_factory=dict)

    # Per-host utilisation (from vHost 'CPU usage %' / 'Memory usage %')
    # Keys are host FQDNs; values are percent (0–100).
    host_cpu_util: dict[str, float] = field(default_factory=dict)
    host_mem_util: dict[str, float] = field(default_factory=dict)

    # Per-VM consumed storage from vPartition (GiB, sum across all partitions)
    vm_partition_consumed_gib: dict[str, float] = field(default_factory=dict)

    # Metadata
    source_file: str = ""
    source_type: str = "rvtools"   # inventory source format identifier
    parse_warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

def _col_index(headers: list, name: str, warn_list: list[str]) -> int | None:
    """Return 0-based column index for a named header, or None on miss."""
    try:
        return headers.index(name)
    except ValueError:
        warn_list.append(f"Column not found: '{name}'")
        return None


def parse(path: str | Path, include_powered_off: bool | None = None) -> RVToolsInventory:
    """
    Parse an RVTools export and return an aggregated RVToolsInventory.

    TCO baseline scope (num_vms, vCPU, memory, storage, Windows/ESU pCores)
    is auto-detected from vHost tab availability when include_powered_off=None:

      vHost tab present  → powered-on VMs only (default).  The vHost tab
        confirms the physical host inventory; powered-off VMs represent idle
        capacity not consuming active hardware resources.
        Override: pass include_powered_off=True to also count powered-off VMs.

      vHost tab absent   → all VMs, powered-on + powered-off (default).
        Without host data the inventory may be incomplete, so every VM is
        included to avoid understating the baseline.
        Override: pass include_powered_off=False to limit to powered-on only.

    Azure migration sizing fields (num_vms_poweredon, total_vcpu_poweredon,
    total_vmemory_gb_poweredon, total_storage_poweredon_gb) are ALWAYS
    powered-on only regardless of this setting.

    Args:
        path: Path to the RVTools .xlsx file.
        include_powered_off: Override for TCO baseline scope.
            None (default) = auto-detect from vHost availability.
            True  = include powered-off VMs (even if vHost is present).
            False = exclude powered-off VMs (even if vHost is absent).
    """
    path = Path(path)
    inv = RVToolsInventory(source_file=str(path))
    warnings = inv.parse_warnings

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    # Counters populated in vInfo block; scope resolved into inv fields after
    # vHost tab availability is determined.
    num_vms_all = 0;       total_vcpu_all = 0;    total_mem_mb_all = 0.0
    total_stor_mib_all = 0.0
    win_vcpus_all = 0;     win_esu_vcpus_all = 0; win_unversioned_all = 0
    num_vms_on = 0;        total_vcpu_on = 0;     total_mem_mb_on = 0.0
    total_stor_mib_on = 0.0
    win_vcpus_on = 0;      win_esu_vcpus_on = 0;  win_unversioned_on = 0
    vhost_found = False
    # SQL-from-Application tracking (all-VM scope, silent columns)
    sql_vcpus_all = 0; sql_vms_all = 0; sql_prod_all = 0; sql_nonprod_all = 0
    sql_env_tagged_count = 0   # SQL VMs that had any non-empty Environment value
    env_all_tagged_count  = 0  # all VMs with any non-empty Environment value
    # Per-VM rightsizing structures (powered-on only)
    _vm_records_on: list[VMRecord] = []
    _vm_to_host: dict[str, str] = {}

    # ------------------------------------------------------------------
    # vInfo tab
    # ------------------------------------------------------------------
    if "vInfo" not in wb.sheetnames:
        warnings.append("vInfo tab not found — no VM data extracted")
    else:
        ws = wb["vInfo"]
        rows = ws.iter_rows(values_only=True)
        headers = list(next(rows))

        ci_name     = _col_index(headers, COL_VM_NAME,    warnings)
        ci_power    = _col_index(headers, COL_POWERSTATE, warnings)
        ci_template = _col_index(headers, COL_TEMPLATE,   [])
        ci_cpu      = _col_index(headers, COL_CPUS,       warnings)
        ci_mem      = _col_index(headers, COL_MEMORY_MB,  warnings)
        ci_stor     = _col_index(headers, COL_IN_USE_MIB, warnings)
        ci_os_cfg   = _col_index(headers, COL_OS_CONFIG,  warnings)
        ci_os_tools = _col_index(headers, COL_OS_TOOLS,   warnings)
        # Optional custom-attribute columns — silent on miss (customer-specific)
        ci_app      = _col_index(headers, "Application",  [])
        ci_env      = _col_index(headers, "Environment",  [])
        # Per-VM rightsizing: host name + vInfo fallback storage columns
        ci_host_vi  = _col_index(headers, "Host",             [])
        ci_prov_mib = _col_index(headers, "Provisioned MiB",  [])
        # 'In Use MiB' already mapped via ci_stor above

        for row in rows:
            if ci_name is not None and row[ci_name] is None:
                continue

            # Skip template entries — they inflate VM counts and licence metrics
            if ci_template is not None and row[ci_template] is True:
                continue

            is_on = (
                ci_power is not None
                and str(row[ci_power] or "").lower() == "poweredon"
            )

            cpus = row[ci_cpu] if ci_cpu is not None else None
            vm_cpus = int(cpus) if isinstance(cpus, (int, float)) else 0
            mem  = row[ci_mem]  if ci_mem  is not None else None
            stor = row[ci_stor] if ci_stor is not None else None

            # OS classification: config-file column is primary (more complete
            # version strings); VMware Tools column is the fallback.
            os_cfg   = str(row[ci_os_cfg]   or "") if ci_os_cfg   is not None else ""
            os_tools = str(row[ci_os_tools] or "") if ci_os_tools is not None else ""

            is_win = (
                bool(_WINDOWS_PATTERN.search(os_cfg))
                or bool(_WINDOWS_PATTERN.search(os_tools))
            )
            is_esu = (
                bool(_WINDOWS_ESU_PATTERN.search(os_cfg))
                or bool(_WINDOWS_ESU_PATTERN.search(os_tools))
            )
            # Versioned = any 4-digit year present in either OS column.
            # Unversioned = Windows Server detected but no year → may be pre-2016.
            is_win_versioned = (
                bool(_WINDOWS_VERSIONED_PATTERN.search(os_cfg))
                or bool(_WINDOWS_VERSIONED_PATTERN.search(os_tools))
            )

            # SQL detection: Application custom attribute (preferred) + OS fallback
            app_str = str(row[ci_app] or "").lower() if ci_app is not None else ""
            env_str = str(row[ci_env] or "").lower() if ci_env is not None else ""
            is_sql = (
                "sql" in app_str
                or "sql server" in os_cfg.lower()
                or "sql server" in os_tools.lower()
            )

            # ── All-VM accumulators ──
            num_vms_all += 1
            if isinstance(cpus, (int, float)):
                total_vcpu_all += vm_cpus
            if isinstance(mem, (int, float)):
                total_mem_mb_all += mem
            if isinstance(stor, (int, float)):
                total_stor_mib_all += stor
            if is_win:
                win_vcpus_all += vm_cpus
                if is_esu:
                    win_esu_vcpus_all += vm_cpus
                elif not is_win_versioned:
                    # No 4-digit year in either OS column — truly unversioned
                    win_unversioned_all += 1
            # Track whether any VM has an environment tag at all
            if env_str:
                env_all_tagged_count += 1

            if is_sql:
                sql_vcpus_all += vm_cpus
                sql_vms_all += 1
                if env_str:
                    sql_env_tagged_count += 1
                    # Explicit non-prod tag → non-production; everything else → production
                    if _ENV_NONPROD_PATTERN.search(env_str):
                        sql_nonprod_all += 1
                    else:
                        sql_prod_all += 1
                else:
                    # No environment tag — default assumption: production
                    sql_prod_all += 1

            # ── Powered-on accumulators ──
            if is_on:
                num_vms_on += 1
                if isinstance(cpus, (int, float)):
                    total_vcpu_on += vm_cpus
                if isinstance(mem, (int, float)):
                    total_mem_mb_on += mem
                if isinstance(stor, (int, float)):
                    total_stor_mib_on += stor
                if is_win:
                    win_vcpus_on += vm_cpus
                    if is_esu:
                        win_esu_vcpus_on += vm_cpus
                    elif not is_win_versioned:
                        win_unversioned_on += 1

                # ── Per-VM record for rightsizing ──
                vm_name_str = str(row[ci_name]) if ci_name is not None else ""
                mem_mib_int = int(mem) if isinstance(mem, (int, float)) else 0
                inuse_gib   = round(float(stor) * MIB_TO_GIB, 4) if isinstance(stor, (int, float)) else 0.0
                prov_mib    = row[ci_prov_mib] if ci_prov_mib is not None else None
                prov_gib    = round(float(prov_mib) * MIB_TO_GIB, 4) if isinstance(prov_mib, (int, float)) else 0.0
                host_name   = str(row[ci_host_vi] or "") if ci_host_vi is not None else ""
                app_str_raw = str(row[ci_app] or "").lower() if ci_app is not None else ""
                # Detect SQL for this VM specifically
                vm_is_sql = (
                    "sql" in app_str_raw
                    or "sql server" in os_cfg.lower()
                    or "sql server" in os_tools.lower()
                )
                _vm_records_on.append(VMRecord(
                    name=vm_name_str,
                    vcpu=vm_cpus,
                    memory_mib=mem_mib_int,
                    host_name=host_name,
                    os_cfg=os_cfg,
                    os_tools=os_tools,
                    app_str=app_str_raw,
                    is_windows=is_win,
                    is_esu=is_esu,
                    is_sql=vm_is_sql,
                    inuse_gib=inuse_gib,
                    provisioned_gib=prov_gib,
                ))
                if vm_name_str and host_name:
                    _vm_to_host[vm_name_str] = host_name

        # Counters ready; scope resolved + inv fields assigned after vHost below.

    # ------------------------------------------------------------------
    # vHost tab
    # ------------------------------------------------------------------
    if "vHost" not in wb.sheetnames:
        warnings.append("vHost tab not found — no host data extracted")
    else:
        vhost_found = True
        ws2 = wb["vHost"]
        rows2 = ws2.iter_rows(values_only=True)
        headers2 = list(next(rows2))

        ci_host = _col_index(headers2, COL_HOST, warnings)
        ci_cores = _col_index(headers2, COL_CORES, warnings)
        ci_hmem = _col_index(headers2, COL_HOST_MEMORY_MB, warnings)
        ci_vpc = _col_index(headers2, COL_VCPUS_PER_CORE, warnings)
        # Host utilisation columns (silent miss — not always exported)
        ci_cpu_pct = _col_index(headers2, "CPU usage %",    [])
        ci_mem_pct = _col_index(headers2, "Memory usage %", [])

        # Region-evidence column indices (silent — no warning on miss)
        ci_dc     = _col_index(headers2, "Datacenter",     [])
        ci_tz     = _col_index(headers2, "Time Zone Name", [])
        ci_gmt    = _col_index(headers2, "GMT Offset",     [])
        ci_domain = _col_index(headers2, "Domain",         [])

        total_cores = 0
        total_hmem_mb = 0.0
        vcpu_per_core_values: list[float] = []
        num_hosts = 0
        dc_names: set[str] = set()
        dc_counts: dict[str, int] = {}
        tz_names: set[str] = set()
        gmt_vals: set[str] = set()
        domain_vals: set[str] = set()
        host_cpu_util: dict[str, float] = {}
        host_mem_util: dict[str, float] = {}

        for row2 in rows2:
            if ci_host is not None and row2[ci_host] is None:
                continue
            host_fqdn = str(row2[ci_host]).strip() if ci_host is not None else ""
            num_hosts += 1

            cores = row2[ci_cores] if ci_cores is not None else None
            if isinstance(cores, (int, float)):
                total_cores += int(cores)

            hmem = row2[ci_hmem] if ci_hmem is not None else None
            if isinstance(hmem, (int, float)):
                total_hmem_mb += hmem

            vpc = row2[ci_vpc] if ci_vpc is not None else None
            if isinstance(vpc, (int, float)) and vpc > 0:
                vcpu_per_core_values.append(float(vpc))

            # Host-level utilisation (percent integers, e.g. 4, 26)
            if host_fqdn:
                cpu_pct = row2[ci_cpu_pct] if ci_cpu_pct is not None else None
                mem_pct = row2[ci_mem_pct] if ci_mem_pct is not None else None
                if isinstance(cpu_pct, (int, float)) and cpu_pct > 0:
                    host_cpu_util[host_fqdn] = float(cpu_pct)
                if isinstance(mem_pct, (int, float)) and mem_pct > 0:
                    host_mem_util[host_fqdn] = float(mem_pct)

            if ci_dc is not None and row2[ci_dc]:
                dc = str(row2[ci_dc]).strip()
                dc_names.add(dc)
                dc_counts[dc] = dc_counts.get(dc, 0) + 1
            if ci_tz is not None and row2[ci_tz]:
                tz_names.add(str(row2[ci_tz]).strip())
            if ci_gmt is not None and row2[ci_gmt] is not None:
                gmt_vals.add(str(row2[ci_gmt]).strip())
            if ci_domain is not None and row2[ci_domain]:
                domain_vals.add(str(row2[ci_domain]).strip().lower())

        inv.num_hosts = num_hosts
        inv.total_host_pcores = total_cores
        inv.total_host_memory_gb = round(total_hmem_mb * MB_TO_GB, 2)
        inv.vcpu_per_core_ratio = (
            round(sum(vcpu_per_core_values) / len(vcpu_per_core_values), 4)
            if vcpu_per_core_values else 1.0
        )
        inv.datacenter_names      = sorted(dc_names)
        inv.datacenter_host_counts = dc_counts
        inv.timezone_names         = sorted(tz_names)
        inv.gmt_offsets            = sorted(gmt_vals)
        inv.domain_names           = sorted(domain_vals)
        inv.host_cpu_util          = host_cpu_util
        inv.host_mem_util          = host_mem_util
        if host_cpu_util:
            _log.debug(f"[rvtools_parser] Host CPU util: {len(host_cpu_util)} hosts with data")
        if host_mem_util:
            _log.debug(f"[rvtools_parser] Host mem util: {len(host_mem_util)} hosts with data")

    # ------------------------------------------------------------------
    # Resolve TCO baseline scope and populate primary inv fields
    # ------------------------------------------------------------------
    inv.vhost_available = vhost_found
    if include_powered_off is None:
        # Auto-detect: powered-on only when vHost confirms complete inventory;
        # all VMs when vHost is absent.
        resolved_poff = not vhost_found
    else:
        resolved_poff = include_powered_off
    inv.include_powered_off_applied = resolved_poff

    if resolved_poff:
        # TCO baseline = all VMs (vHost absent, or explicit override=True)
        inv.num_vms               = num_vms_all
        inv.total_vcpu            = total_vcpu_all
        inv.total_vmemory_gb      = round(total_mem_mb_all * MB_TO_GB, 2)
        inv.total_storage_in_use_gb = round(total_stor_mib_all * MIB_TO_GB, 2)
        inv._win_vcpus     = win_vcpus_all      # type: ignore[attr-defined]
        inv._win_esu_vcpus = win_esu_vcpus_all  # type: ignore[attr-defined]
        win_unversioned    = win_unversioned_all
    else:
        # TCO baseline = powered-on only (vHost present, or explicit override=False)
        inv.num_vms               = num_vms_on
        inv.total_vcpu            = total_vcpu_on
        inv.total_vmemory_gb      = round(total_mem_mb_on * MB_TO_GB, 2)
        inv.total_storage_in_use_gb = round(total_stor_mib_on * MIB_TO_GB, 2)
        inv._win_vcpus     = win_vcpus_on       # type: ignore[attr-defined]
        inv._win_esu_vcpus = win_esu_vcpus_on   # type: ignore[attr-defined]
        win_unversioned    = win_unversioned_on

    # Azure sizing: always powered-on regardless of TCO scope
    inv.num_vms_poweredon          = num_vms_on
    inv.total_vcpu_poweredon       = total_vcpu_on
    inv.total_vmemory_gb_poweredon = round(total_mem_mb_on * MB_TO_GB, 2)
    inv.total_storage_poweredon_gb = round(total_stor_mib_on * MIB_TO_GB, 2)

    # Store per-VM records and host mapping
    inv.vm_records  = _vm_records_on
    inv.vm_to_host  = _vm_to_host

    inv.windows_vms_unknown_version  = win_unversioned
    inv.esu_count_may_be_understated = win_unversioned > 0

    # Log which scope was applied and why
    scope_label = "all VMs" if resolved_poff else "powered-on VMs only"
    if include_powered_off is not None:
        reason = f"override (include_powered_off={include_powered_off})"
    elif vhost_found:
        reason = "vHost tab present"
    else:
        reason = "vHost tab absent"
    _log.debug(f"[rvtools_parser] TCO baseline: {scope_label} ({inv.num_vms:,} VMs) — {reason}")

    if win_unversioned > 0:
        warnings.append(
            f"ESU undercount likely: {win_unversioned} Windows Server VMs in the "
            f"TCO baseline scope have no version string in either OS column "
            f"(config file or VMware Tools).  These are likely pre-2016 VMs "
            f"(2003/2008/2008 R2) that are ESU-eligible but undetectable from "
            f"RVtools OS strings alone.  Review and override 'pCores with ESU' "
            f"in the intake form using a separate OS audit (e.g. MAP Toolkit, "
            f"Azure Migrate, or manual review)."
        )

    # ------------------------------------------------------------------
    # Derive Windows/SQL pCore estimates using the vCPU/core ratio
    # ------------------------------------------------------------------
    ratio = inv.vcpu_per_core_ratio if inv.vcpu_per_core_ratio > 0 else 1.0
    win_vcpus = getattr(inv, "_win_vcpus", 0)
    win_esu_vcpus = getattr(inv, "_win_esu_vcpus", 0)
    inv.pcores_with_windows_server = round(win_vcpus / ratio)
    inv.pcores_with_windows_esu = round(win_esu_vcpus / ratio)

    # SQL pCores: prefer Application-detected count; fall back to 10% of Windows
    inv.sql_vms_detected     = sql_vms_all
    inv.sql_vms_prod         = sql_prod_all
    inv.sql_vms_nonprod      = sql_nonprod_all
    inv.sql_prod_assumed     = (sql_vms_all > 0 and sql_env_tagged_count == 0)
    inv.env_tagging_present  = env_all_tagged_count > 0
    if sql_vcpus_all > 0:
        inv.pcores_with_sql_server = round(sql_vcpus_all / ratio)
        inv.sql_detection_source   = "application"
        # ESU SQL pCores: proportion of SQL pCores relative to total Windows
        win_pcore_total = max(inv.pcores_with_windows_server, 1)
        sql_esu_fraction = inv.pcores_with_windows_esu / win_pcore_total
        inv.pcores_with_sql_esu = round(inv.pcores_with_sql_server * sql_esu_fraction)
        prod_note = "assumed prod — no env tags" if inv.sql_prod_assumed else f"{sql_prod_all} Prod / {sql_nonprod_all} non-Prod"
        _log.debug(
            f"[rvtools_parser] SQL detection (Application): "
            f"{sql_vms_all} VMs → {inv.pcores_with_sql_server} pCores  ({prod_note})"
        )
    else:
        inv.pcores_with_sql_server = round(inv.pcores_with_windows_server * 0.10)
        inv.pcores_with_sql_esu    = round(inv.pcores_with_windows_esu * 0.10)
        inv.sql_detection_source   = "default"

    # Clean up private attrs
    for attr in ("_win_vcpus", "_win_esu_vcpus"):
        if hasattr(inv, attr):
            delattr(inv, attr)

    # ------------------------------------------------------------------
    # vCPU tab — per-VM CPU utilisation (Overall MHz / Max MHz)
    # Builds vm_cpu_util dict AND fleet P95 for summary display.
    # Powered-on VMs only; VMs with Max == 0 are skipped.
    # ------------------------------------------------------------------
    if "vCPU" in wb.sheetnames:
        ws_cpu = wb["vCPU"]
        rc = ws_cpu.iter_rows(values_only=True)
        hc = list(next(rc))
        ci_cp  = _col_index(hc, "Powerstate", [])
        ci_vm_cpu = _col_index(hc, "VM",       [])
        ci_max = _col_index(hc, "Max",        [])
        ci_ovr = _col_index(hc, "Overall",    [])
        cpu_utils: list[float] = []
        vm_cpu_util: dict[str, float] = {}
        for row in rc:
            if ci_cp is None:
                break
            if str(row[ci_cp] or "").lower() != "poweredon":
                continue
            mx = row[ci_max] if ci_max is not None else None
            ov = row[ci_ovr] if ci_ovr is not None else None
            if isinstance(mx, (int, float)) and isinstance(ov, (int, float)) and mx > 0:
                util = float(ov) / float(mx)
                cpu_utils.append(util)
                vm_name_cpu = str(row[ci_vm_cpu]) if ci_vm_cpu is not None and row[ci_vm_cpu] else ""
                if vm_name_cpu:
                    vm_cpu_util[vm_name_cpu] = round(util, 4)
        if cpu_utils:
            cpu_utils.sort()
            p = min(int(len(cpu_utils) * 0.95), len(cpu_utils) - 1)
            inv.cpu_util_p95 = round(cpu_utils[p], 4)
            inv.cpu_util_p95_vm_count = len(cpu_utils)
            inv.vm_cpu_util = vm_cpu_util
            _log.debug(
                f"[rvtools_parser] CPU P95 utilisation: {inv.cpu_util_p95:.1%}"
                f" ({inv.cpu_util_p95_vm_count:,} powered-on VMs)"
            )
    else:
        warnings.append(
            "vCPU tab not found — per-VM CPU utilisation unavailable. "
            "vHost CPU usage % will be used as proxy where available; "
            "otherwise fallback factor (retain 40% vCPU) applies."
        )

    # ------------------------------------------------------------------
    # vMemory tab — per-VM memory utilisation (Consumed MiB / Size MiB)
    # Builds vm_mem_util dict AND fleet P95 for summary display.
    # ------------------------------------------------------------------
    if "vMemory" in wb.sheetnames:
        ws_mem = wb["vMemory"]
        rm = ws_mem.iter_rows(values_only=True)
        hm = list(next(rm))
        ci_mp     = _col_index(hm, "Powerstate", [])
        ci_vm_mem = _col_index(hm, "VM",         [])
        ci_sz     = _col_index(hm, "Size MiB",   [])
        ci_con    = _col_index(hm, "Consumed",    [])
        mem_utils: list[float] = []
        vm_mem_util: dict[str, float] = {}
        for row in rm:
            if ci_mp is None:
                break
            if str(row[ci_mp] or "").lower() != "poweredon":
                continue
            sz = row[ci_sz]  if ci_sz  is not None else None
            co = row[ci_con] if ci_con is not None else None
            if isinstance(sz, (int, float)) and isinstance(co, (int, float)) and sz > 0:
                util = float(co) / float(sz)
                mem_utils.append(util)
                vm_name_mem = str(row[ci_vm_mem]) if ci_vm_mem is not None and row[ci_vm_mem] else ""
                if vm_name_mem:
                    vm_mem_util[vm_name_mem] = round(util, 4)
        if mem_utils:
            mem_utils.sort()
            p = min(int(len(mem_utils) * 0.95), len(mem_utils) - 1)
            inv.memory_util_p95 = round(mem_utils[p], 4)
            inv.memory_util_p95_vm_count = len(mem_utils)
            inv.vm_mem_util = vm_mem_util
            _log.debug(
                f"[rvtools_parser] Memory P95 utilisation: {inv.memory_util_p95:.1%}"
                f" ({inv.memory_util_p95_vm_count:,} powered-on VMs)"
            )
    else:
        warnings.append(
            "vMemory tab not found — per-VM memory utilisation unavailable. "
            "vHost Memory usage % will be used as proxy where available; "
            "otherwise fallback factor (retain 60% memory) applies."
        )

    # ------------------------------------------------------------------
    # vDisk tab — provisioned disk capacity per VM
    # Azure managed disk cost is based on provisioned tier size, not used
    # bytes, so this is the correct basis for the Azure storage estimate.
    # ------------------------------------------------------------------
    if "vDisk" in wb.sheetnames:
        ws_disk = wb["vDisk"]
        rd = ws_disk.iter_rows(values_only=True)
        hd = list(next(rd))
        ci_dvm = _col_index(hd, "VM",           [])
        ci_dp  = _col_index(hd, "Powerstate",   [])
        ci_cap = _col_index(hd, "Capacity MiB", [])
        ci_dt  = _col_index(hd, "Template",     [])
        total_cap_mib_all = 0.0
        total_cap_mib_on  = 0.0
        vm_disks: dict[str, list[float]] = {}
        for row in rd:
            if ci_dt is not None and row[ci_dt] is True:
                continue
            cap = row[ci_cap] if ci_cap is not None else None
            if not isinstance(cap, (int, float)):
                continue
            total_cap_mib_all += float(cap)
            is_on = (
                ci_dp is not None
                and str(row[ci_dp] or "").lower() == "poweredon"
            )
            if is_on:
                total_cap_mib_on += float(cap)
                if ci_dvm is not None and row[ci_dvm]:
                    vm_name = str(row[ci_dvm])
                    if vm_name not in vm_disks:
                        vm_disks[vm_name] = []
                    vm_disks[vm_name].append(round(float(cap) * MIB_TO_GB, 4))
        inv.total_disk_provisioned_gb           = round(total_cap_mib_all * MIB_TO_GB, 2)
        inv.total_disk_provisioned_poweredon_gb = round(total_cap_mib_on  * MIB_TO_GB, 2)
        inv.vm_disk_sizes_gb = vm_disks

        # Also populate VMRecord.disk_sizes_gib (GiB = MiB ÷ 1024)
        # Build a lookup: vm_name → [gib, gib, ...]
        vm_disks_gib: dict[str, list[float]] = {}
        for vm_name, sizes_dec_gb in vm_disks.items():
            # sizes_dec_gb was stored as MIB_TO_GB; recalculate as GiB from scratch
            # We need raw MiB values — re-compute: GiB = dec_GB × 953.67 / 1024
            # Simpler: store GiB directly by re-reading; but to avoid a second file pass
            # we convert: dec_gb × (953.67 / 1024) = GiB
            vm_disks_gib[vm_name] = [round(v * 953.67 / 1024.0, 4) for v in sizes_dec_gb]
        for vm_rec in inv.vm_records:
            if vm_rec.name in vm_disks_gib:
                vm_rec.disk_sizes_gib = vm_disks_gib[vm_rec.name]

        _log.debug(
            f"[rvtools_parser] vDisk provisioned: all={inv.total_disk_provisioned_gb:,.0f} GB  "
            f"powered-on={inv.total_disk_provisioned_poweredon_gb:,.0f} GB  "
            f"({len(vm_disks):,} VMs, {sum(len(v) for v in vm_disks.values()):,} disks)"
        )

    # ------------------------------------------------------------------
    # vPartition tab — per-VM consumed filesystem storage (GiB)
    # Used as storage fallback when vDisk tab is absent.
    # ------------------------------------------------------------------
    if "vPartition" in wb.sheetnames:
        ws_part = wb["vPartition"]
        rp = ws_part.iter_rows(values_only=True)
        hp = list(next(rp))
        ci_pvm  = _col_index(hp, "VM",           [])
        ci_ppow = _col_index(hp, "Powerstate",   [])
        ci_pcon = _col_index(hp, "Consumed MiB", [])
        part_consumed: dict[str, float] = {}
        for row in rp:
            if ci_ppow is not None and str(row[ci_ppow] or "").lower() != "poweredon":
                continue
            vm_n = str(row[ci_pvm] or "") if ci_pvm is not None else ""
            con  = row[ci_pcon] if ci_pcon is not None else None
            if vm_n and isinstance(con, (int, float)) and con > 0:
                part_consumed[vm_n] = part_consumed.get(vm_n, 0.0) + float(con) * MIB_TO_GIB
        if part_consumed:
            inv.vm_partition_consumed_gib = {k: round(v, 4) for k, v in part_consumed.items()}
            # Populate VMRecord.partition_consumed_gib
            for vm_rec in inv.vm_records:
                if vm_rec.name in part_consumed:
                    vm_rec.partition_consumed_gib = round(part_consumed[vm_rec.name], 4)
            _log.debug(
                f"[rvtools_parser] vPartition consumed: {len(part_consumed):,} VMs, "
                f"total {sum(part_consumed.values()):,.0f} GiB"
            )

    # ------------------------------------------------------------------
    # vMetaData tab — vCenter FQDN(s) for region inference
    # ------------------------------------------------------------------
    if "vMetaData" in wb.sheetnames:
        ws_md = wb["vMetaData"]
        rmd = ws_md.iter_rows(values_only=True)
        hmd = list(next(rmd))
        ci_srv = _col_index(hmd, "Server", [])
        fqdns: list[str] = []
        for row in rmd:
            if ci_srv is not None and row[ci_srv]:
                fqdns.append(str(row[ci_srv]).strip().lower())
        inv.vcenter_fqdns = fqdns

    if warnings:
        _log.debug(f"[rvtools_parser] {len(warnings)} warning(s):")
        for w in warnings:
            _log.debug(f"  - {w}")

    return inv


def summarize(inv: RVToolsInventory) -> None:
    """Print a human-readable summary of a parsed RVToolsInventory."""
    scope = "all VMs" if inv.include_powered_off_applied else "powered-on only"
    _log.debug(f"Source: {inv.source_file}")
    _log.debug(f"  vHost data available:         {'yes' if inv.vhost_available else 'no':>10}")
    _log.debug(f"  TCO baseline scope:           {scope}")
    _log.debug()
    _log.debug(f"  ── On-Prem TCO Baseline ({scope}) ──")
    _log.debug(f"  VMs:                          {inv.num_vms:>10,}")
    _log.debug(f"  Hosts:                        {inv.num_hosts:>10,}")
    _log.debug(f"  Total vCPU:                   {inv.total_vcpu:>10,}")
    _log.debug(f"  Total vMemory (GB):           {inv.total_vmemory_gb:>10,.1f}")
    _log.debug(f"  Storage in use (GB):          {inv.total_storage_in_use_gb:>10,.1f}")
    _log.debug(f"  Host pCores (total):          {inv.total_host_pcores:>10,}")
    _log.debug(f"  Host Memory GB:               {inv.total_host_memory_gb:>10,.1f}")
    _log.debug(f"  vCPUs per pCore (avg):        {inv.vcpu_per_core_ratio:>10.3f}")
    _log.debug(f"  pCores w/ Win Server:         {inv.pcores_with_windows_server:>10,}")
    esu_note = "  ⚑ may be understated" if inv.esu_count_may_be_understated else ""
    _log.debug(f"  pCores w/ Win ESU:            {inv.pcores_with_windows_esu:>10,}{esu_note}")
    if inv.esu_count_may_be_understated:
        _log.debug(f"  Windows VMs w/ unknown ver:   {inv.windows_vms_unknown_version:>10,}  (check OS audit)")
    _log.debug(f"  pCores w/ SQL Server:         {inv.pcores_with_sql_server:>10,}")
    _log.debug(f"  pCores w/ SQL ESU:            {inv.pcores_with_sql_esu:>10,}")
    _log.debug()
    _log.debug("  ── Azure Migration Target (powered-on only) ──")
    _log.debug(f"  VMs (powered-on):             {inv.num_vms_poweredon:>10,}")
    _log.debug(f"  vCPU (powered-on):            {inv.total_vcpu_poweredon:>10,}")
    _log.debug(f"  vMemory GB (powered-on):      {inv.total_vmemory_gb_poweredon:>10,.1f}")
    _log.debug(f"  Storage GB (powered-on):      {inv.total_storage_poweredon_gb:>10,.1f}")
    _log.debug()
    _log.debug("  ── Utilisation Telemetry ──")
    if inv.cpu_util_p95 > 0:
        _log.debug(f"  CPU P95 utilisation:          {inv.cpu_util_p95:>9.1%}  ({inv.cpu_util_p95_vm_count:,} VMs)")
    else:
        _log.debug("  CPU P95 utilisation:            n/a  (vCPU tab absent or all VMs powered-off)")
    if inv.memory_util_p95 > 0:
        _log.debug(f"  Memory P95 utilisation:       {inv.memory_util_p95:>9.1%}  ({inv.memory_util_p95_vm_count:,} VMs)")
    else:
        _log.debug("  Memory P95 utilisation:         n/a  (vMemory tab absent)")
    _log.debug()
    _log.debug("  ── Region Evidence ──")
    if inv.datacenter_names:
        _log.debug(f"  Datacenter(s):  {', '.join(inv.datacenter_names)}")
    if inv.timezone_names:
        _log.debug(f"  Time zone(s):   {', '.join(inv.timezone_names)}")
    if inv.gmt_offsets:
        _log.debug(f"  GMT offset(s):  {', '.join(inv.gmt_offsets)}")
    if inv.domain_names:
        _log.debug(f"  Domain(s):      {', '.join(inv.domain_names)}")
    if inv.vcenter_fqdns:
        _log.debug(f"  vCenter FQDN(s): {', '.join(inv.vcenter_fqdns)}")
    if inv.parse_warnings:
        _log.debug(f"\n  Warnings ({len(inv.parse_warnings)}):")
        for w in inv.parse_warnings:
            _log.debug(f"    \u26a0  {w}")
